import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import xlrd, datetime, io, zipfile, base64, os
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.colors import white, black

# Load Mt Fuji icon for page/PWA icon
def _load_app_icon():
    try:
        from PIL import Image
        p = os.path.join(os.path.dirname(__file__), 'icon.b64')
        with open(p, 'r') as f:
            return Image.open(io.BytesIO(base64.b64decode(f.read())))
    except Exception:
        return "🗻"

st.set_page_config(page_title="Hotel Guest Processor", page_icon=_load_app_icon(), layout="centered")

# ── Load embedded templates ──────────────────────────────────────────────
@st.cache_resource
def load_template(name):
    path = os.path.join(os.path.dirname(__file__), f'tmpl_{name}.b64')
    with open(path, 'r') as f:
        return base64.b64decode(f.read())

# ── Lookup tables ─────────────────────────────────────────────────────────
# Full nationality mapping (normalized keys → "CODE - Name") — 350 entries
import unicodedata as _ud, re as _re
def _norm_nat(s):
    s = str(s).lower().strip()
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return _re.sub(r'[^a-z0-9]', '', s)

NAT_NORM = {
    "achentina": "ARG - Argentina",
    "acmenia": "ARM - Armenia",
    "adecbaigian": "AZE - Azerbaijan",
    "aicap": "EGY - Egypt",
    "airolen": "IRL - Ireland",
    "aixolen": "ISL - Iceland",
    "albania": "ALB - Albania",
    "anbani": "ALB - Albania",
    "andorra": "AND - Andorra",
    "anggola": "AGO - Angola",
    "angola": "AGO - Angola",
    "anh": "GBR - United Kingdom",
    "anmach": "DNK - Denmark",
    "ano": "IND - India",
    "ao": "AUT - Austria",
    "aosip": "CYP - Cyprus",
    "arapthongnhat": "ARE - United Arab Emirates",
    "arapxaui": "SAU - Saudi Arabia",
    "argentina": "ARG - Argentina",
    "armenia": "ARM - Armenia",
    "australia": "AUS - Australia",
    "austria": "AUT - Austria",
    "azerbaijan": "AZE - Azerbaijan",
    "bacbaot": "BRB - Barbados",
    "bahama": "BHS - Bahamas",
    "bahamas": "BHS - Bahamas",
    "bahrain": "BHR - Bahrain",
    "balan": "POL - Poland",
    "bangladesh": "BGD - Bangladesh",
    "banglaet": "BGD - Bangladesh",
    "barain": "BHR - Bahrain",
    "barbados": "BRB - Barbados",
    "becmua": "BMU - Bermuda",
    "belarus": "BLR - Belarus",
    "belarut": "BLR - Belarus",
    "belgium": "BEL - Belgium",
    "belixe": "BLZ - Belize",
    "belize": "BLZ - Belize",
    "benanh": "BEN - Benin",
    "benin": "BEN - Benin",
    "bermuda": "BMU - Bermuda",
    "bhutan": "BTN - Bhutan",
    "bi": "BEL - Belgium",
    "boaonha": "PRT - Portugal",
    "bolivia": "BOL - Bolivia",
    "bosniaandherzegovina": "BIH - Bosnia and Herzegovina",
    "botswana": "BWA - Botswana",
    "botxoana": "BWA - Botswana",
    "boxniahecdegovina": "BIH - Bosnia and Herzegovina",
    "bradin": "BRA - Brazil",
    "brazil": "BRA - Brazil",
    "britishindiaoceanterritory": "IOT - British India Ocean Territory",
    "bulgaria": "BGR - Bulgaria",
    "bungari": "BGR - Bulgaria",
    "buockinaphaxo": "BFA - Burkina Faso",
    "burkinafaso": "BFA - Burkina Faso",
    "burundi": "BDI - Burundi",
    "buruni": "BDI - Burundi",
    "butan": "BTN - Bhutan",
    "cameroon": "CMR - Cameroon",
    "camorun": "CMR - Cameroon",
    "canada": "CAN - Canada",
    "capeverde": "CPV - Cape Verde",
    "capve": "CPV - Cape Verde",
    "chad": "TCD - Chad",
    "charapxyri": "SYR - Syrian Arab Republic",
    "chdcndtrieutien": "PRK - Korea Democratic Peoples Republic of",
    "chhanquoc": "KOR - Korea (South)",
    "chhoigiaoiran": "IRN - Iran Ilasmic Republic of",
    "chile": "CHL - Chile",
    "china": "CHN - China",
    "chinataiwan": "CHN - China",
    "chlienbanguc": "D - Germany",
    "chmaxeonia": "MKD - Macedonia",
    "chominicana": "DMA - Dominica",
    "colombia": "COL - Colombia",
    "como": "COM - Comoros",
    "comoros": "COM - Comoros",
    "conggo": "COG - Congo",
    "conghoasec": "CZE - Czech Republic",
    "congo": "COG - Congo",
    "congquocanora": "AND - Andorra",
    "congquoclichtenxten": "LIE - Liechtenstein",
    "congquocmonaco": "MCO - Monaco",
    "cooet": "KWT - Kuwait",
    "costarica": "CRI - Costa Rica",
    "cotedivoire": "CIV - Cote d' Ivoire",
    "cotivoa": "CIV - Cote d' Ivoire",
    "coxtarica": "CRI - Costa Rica",
    "croatia": "HRV - Croatia",
    "cuba": "CUB - Cuba",
    "cyprus": "CYP - Cyprus",
    "czechrepublic": "CZE - Czech Republic",
    "dambia": "ZMB - Zambia",
    "denmark": "DNK - Denmark",
    "dimbabue": "ZWE - Zimbabwe",
    "djibouti": "DJI - Djibouti",
    "dominica": "DMA - Dominica",
    "dominicana": "DMA - Dominica",
    "ecuador": "ECU - Ecuador",
    "ecuao": "ECU - Ecuador",
    "egypt": "EGY - Egypt",
    "elsalvador": "SLV - El Salvado",
    "enxanvao": "SLV - El Salvado",
    "equatorialguinea": "GNQ - Equatorial Guinea",
    "eritoria": "ERI - Eritrea",
    "eritrea": "ERI - Eritrea",
    "estonia": "EST - Estonia",
    "ethiopia": "ETH - Ethiopia",
    "etiopia": "ETH - Ethiopia",
    "extonia": "EST - Estonia",
    "fiji": "FJI - Fiji",
    "finland": "FIN - Finland",
    "france": "FRA - France",
    "francemetropolitan": "FRA - France",
    "gabon": "GAB - Gabon",
    "gabong": "GAB - Gabon",
    "gambia": "GMB - Gambia",
    "gana": "GHA - Ghana",
    "georgia": "GEO - Georgia",
    "germany": "D - Germany",
    "ghana": "GHA - Ghana",
    "ghine": "GIN - Guinea",
    "ghinebitxao": "GNB - Guinea-Bissau",
    "ghinexichao": "GNQ - Equatorial Guinea",
    "giamahiriiaaraplibinhandan": "LBY - Libyan Arab Jamahiriya",
    "gibraltar": "GIB - Gibraltar",
    "gibranta": "GIB - Gibraltar",
    "goatemala": "GTM - Guatemala",
    "greece": "GRC - Greece",
    "greenland": "GRL - Greenland",
    "grenaa": "GRD - Grenada",
    "grenada": "GRD - Grenada",
    "grinlon": "GRL - Greenland",
    "grudia": "GEO - Georgia",
    "guatemala": "GTM - Guatemala",
    "guina": "GUY - Guyana",
    "guinea": "GIN - Guinea",
    "guineabissau": "GNB - Guinea-Bissau",
    "guyana": "GUY - Guyana",
    "haiti": "HTI - Haiti",
    "halan": "NLD - Netherland",
    "hanquoc": "KOR - Korea (South)",
    "honduras": "HND - Honduras",
    "hondurat": "HND - Honduras",
    "hylap": "GRC - Greece",
    "iaphanthuoclienhiepanh": "GBD - United Kingdom British Territories Citizen",
    "ibouti": "DJI - Djibouti",
    "iceland": "ISL - Iceland",
    "india": "IND - India",
    "indonesia": "IDN - Indonesia",
    "inonexia": "IDN - Indonesia",
    "irac": "IRQ - Iraq",
    "iran": "IRN - Iran Ilasmic Republic of",
    "iraq": "IRQ - Iraq",
    "ireland": "IRL - Ireland",
    "israel": "ISR - Israel",
    "italia": "ITA - Italy",
    "italy": "ITA - Italy",
    "ixraen": "ISR - Israel",
    "jamaica": "JAM - Jamaica",
    "japan": "JPN - Japan",
    "jocan": "JOR - Jordan",
    "jordan": "JOR - Jordan",
    "kadacxtan": "KAZ - Kazakhstan",
    "kazakhstan": "KAZ - Kazakhstan",
    "kenia": "KEN - Kenya",
    "kenya": "KEN - Kenya",
    "kiecghidia": "KGZ - Kyrgyzstan",
    "kiribati": "KIR - Kiribati",
    "koreademocraticpeoplesrepublic": "PRK - Korea Democratic Peoples Republic of",
    "koreasouth": "KOR - Korea (South)",
    "kosovo": "RKS - Kosovo",
    "kuwait": "KWT - Kuwait",
    "kyrgyzstan": "KGZ - Kyrgyzstan",
    "latvia": "LVA - Latvia",
    "lebanon": "LBN - Lebanon",
    "lesotho": "LSO - Lesotho",
    "lexotho": "LSO - Lesotho",
    "liban": "LBN - Lebanon",
    "liberia": "LBR - Liberia",
    "libya": "LBY - Libyan Arab Jamahiriya",
    "liechtenstein": "LIE - Liechtenstein",
    "lienbangnga": "RUS - Russia",
    "lithuania": "LTU - Lithuania",
    "luxembourg": "LUX - Luxembourg",
    "luychxembua": "LUX - Luxembourg",
    "maagaxca": "MDG - Madagascar",
    "macedonia": "MKD - Macedonia",
    "madagascar": "MDG - Madagascar",
    "malaixia": "MYS - Malaysia",
    "malauy": "MWI - Malawi",
    "malawi": "MWI - Malawi",
    "malaysia": "MYS - Malaysia",
    "maldives": "MDV - Maldives",
    "mali": "MLI - Mali",
    "malta": "MLT - Malta",
    "manivo": "MDV - Maldives",
    "manta": "MLT - Malta",
    "maroc": "MAR - Morocco",
    "marshallislands": "MHL - Marshall Islands",
    "mauritania": "MRT - Mauritania",
    "mauritius": "MUS - Mauritius",
    "mexico": "MEX - Mexico",
    "mianma": "MMR - Myanmar",
    "micronesia": "FSM - Micronesia",
    "modambich": "MOZ - Mozambique",
    "moldova": "MDA - Moldova",
    "monaco": "MCO - Monaco",
    "mongco": "MNG - Mongolia",
    "mongolia": "MNG - Mongolia",
    "monova": "MDA - Moldova",
    "montenegro": "MNE - Montenegro",
    "montserrat": "MSR - Montserrat",
    "monxerat": "MSR - Montserrat",
    "moratani": "MRT - Mauritania",
    "morixo": "MUS - Mauritius",
    "morocco": "MAR - Morocco",
    "mozambique": "MOZ - Mozambique",
    "my": "USA - United States of America",
    "myanmarburma": "MMR - Myanmar",
    "namibia": "NAM - Namibia",
    "nauru": "NRU - Nauru",
    "nauy": "NOR - Norway",
    "nepal": "NPL - Nepal",
    "nepan": "NPL - Nepal",
    "netherland": "NLD - Netherland",
    "netherlandantilles": "NLD - Netherland",
    "newzealand": "NZL - New Zealand",
    "nhatban": "JPN - Japan",
    "nicaragoa": "NIC - Nicaragua",
    "nicaragua": "NIC - Nicaragua",
    "niger": "NER - Niger",
    "nigeria": "NGA - Nigeria",
    "nigie": "NER - Niger",
    "nigieria": "NGA - Nigeria",
    "niudilan": "NZL - New Zealand",
    "norway": "NOR - Norway",
    "oman": "OMN - Oman",
    "ominica": "DMA - Dominica",
    "ongtimo": "TLS - Timor Leste",
    "oxtraylia": "AUS - Australia",
    "pakistan": "PAK - Pakistan",
    "pakixtan": "PAK - Pakistan",
    "palau": "PLW - Palau",
    "palestine": "PSE - Palestine",
    "palextin": "PSE - Palestine",
    "panama": "PAN - Panama",
    "papuanewguinea": "PNG - Papua New Guinea",
    "papuaniughine": "PNG - Papua New Guinea",
    "paragoay": "PRY - Paraguay",
    "paraguay": "PRY - Paraguay",
    "peru": "PER - Peru",
    "phanlan": "FIN - Finland",
    "phap": "FRA - France",
    "philippin": "PHL - Philippines",
    "philippine": "PHL - Philippines",
    "poland": "POL - Poland",
    "portugal": "PRT - Portugal",
    "qatar": "QAT - Qatar",
    "quanaoantithuochalan": "NLD - Netherland",
    "quanaomacsan": "MHL - Marshall Islands",
    "quanaonamgrudiavanamsanuych": "GEO - Georgia",
    "quanaoxaysen": "SYC - Seychelles",
    "quata": "QAT - Qatar",
    "romania": "ROU - Romania",
    "ruana": "RWA - Rwanda",
    "rumani": "ROU - Romania",
    "russia": "RUS - Russia",
    "rwanda": "RWA - Rwanda",
    "saintlucia": "LCA - Saint Lucia",
    "sanmarino": "SMR - San Marino",
    "sat": "TCD - Chad",
    "saudiarabia": "SAU - Saudi Arabia",
    "scotland": "SC- - Scotland",
    "senegal": "SEN - Senegal",
    "serbia": "SRB - Serbia",
    "seychelles": "SYC - Seychelles",
    "singapore": "SGP - Singapore",
    "slovakia": "SVK - Slovakia",
    "slovenia": "SVN - Slovenia",
    "somalia": "SOM - Somalia",
    "southgeorgiaandthesouths": "GEO - Georgia",
    "spain": "ESP - Spain",
    "srilanka": "LKA - Sri Lanka",
    "sudan": "SDN - Sudan",
    "suriname": "SUR - Suriname",
    "swaziland": "SWZ - Swaziland",
    "sweden": "SWE - Sweden",
    "switzerland": "CHE - Switzerland",
    "syria": "SYR - Syrian Arab Republic",
    "tagikixtan": "TJK - Tajikistan",
    "tajikistan": "TJK - Tajikistan",
    "taybannha": "ESP - Spain",
    "thailan": "THA - Thailand",
    "thailand": "THA - Thailand",
    "thonhiky": "TUR - Turkey",
    "thuyien": "SWE - Sweden",
    "thuysi": "CHE - Switzerland",
    "timorleste": "TLS - Timor Leste",
    "tochucdantocthongnhat": "UNO - United Nations Organization",
    "togo": "TGO - Togo",
    "tonga": "TON - Tonga",
    "trungquoc": "CHN - China",
    "trungquocailoan": "CHN - China",
    "tunidi": "TUN - Tunisia",
    "tunisia": "TUN - Tunisia",
    "tuocmenixtan": "TKM - Turkmenistan",
    "turkey": "TUR - Turkey",
    "turkmenistan": "TKM - Turkmenistan",
    "tuvalu": "TUV - Tuvalu",
    "uc": "D - Germany",
    "ucraina": "UKR - Ukraine",
    "udobekixtan": "UZB - Uzbekistan",
    "uganda": "UGA - Uganda",
    "ukraine": "UKR - Ukraine",
    "unitedarabemirates": "ARE - United Arab Emirates",
    "unitedkingdom": "GBD - United Kingdom British Territories Citizen",
    "unitednationsorganization": "UNO - United Nations Organization",
    "unitedstates": "USA - United States of America",
    "urugoay": "URY - Uruguay",
    "uruguay": "URY - Uruguay",
    "uzbekistan": "UZB - Uzbekistan",
    "vanuatu": "VUT - Vanuatu",
    "vaticancity": "VAT - Holy See (Vatican City State )",
    "vaticang": "VAT - Holy See (Vatican City State )",
    "veneduela": "VEN - Venezuela",
    "venezuela": "VEN - Venezuela",
    "vietnam": "VNM - Viet Nam",
    "vungatthuocanhoanoduong": "IOT - British India Ocean Territory",
    "vungthuophap": "FRA - France",
    "vuongquocnauy": "NOR - Norway",
    "westernsamoa": "WSM - Western Samoa",
    "xamoa": "WSM - Western Samoa",
    "xanhluxia": "LCA - Saint Lucia",
    "xanmarino": "SMR - San Marino",
    "xcolent": "SC- - Scotland",
    "xecbia": "SRB - Serbia",
    "xenegan": "SEN - Senegal",
    "xingapo": "SGP - Singapore",
    "xlovakia": "SVK - Slovakia",
    "xoadilen": "SWZ - Swaziland",
    "xomali": "SOM - Somalia",
    "xrilanca": "LKA - Sri Lanka",
    "xuang": "SDN - Sudan",
    "xurinam": "SUR - Suriname",
    "y": "ITA - Italy",
    "yemen": "YEM - Yemen",
    "zambia": "ZMB - Zambia",
    "zimbabwe": "ZWE - Zimbabwe"
}

def lookup_nat_kbtt(raw):
    """Khớp thông minh: chuẩn hóa dấu/khoảng trắng để tìm mã quốc tịch."""
    if not raw: return ''
    raw = str(raw).strip()
    key = _norm_nat(raw)
    if key in NAT_NORM:
        return NAT_NORM[key]
    # already in CODE - Name form?
    if _re.match(r'^[A-Z]{2,3} - ', raw):
        return raw
    return raw  # unknown -> keep original (sẽ hiện cảnh báo)

NAT_DK14 = {
    'RUS':'Russia  (Liên bang Nga)','UZB':'Uzbekistan  ( U-dơ-bê-ki-xtan )',
    'KAZ':'Kazakhstan  ( Ka-dắc-xtan )','KOR':'Korea (South)  ( CH Hàn Quốc )',
    'KGZ':'Kyrgyzstan  ( Kiếc-ghi-di-a )','TJK':'Tajikistan  ( Ta-gi-ki-xtan )',
    'UKR':'Ukraine  ( U-crai-na )','USA':'United States  ( Mỹ )',
    'VNM':'Vietnam  ( Việt Nam )','CAN':'Canada  ( Ca-na-da )',
    'GBR':'United Kingdom  ( Anh )','AUS':'Australia  ( Ô-xtrây-li-a )',
    'BLR':'Belarus  ( Bê-la-rút )','CHN':'China  ( Trung Quốc )',
    'DEU':'Germany  ( Đức )','MDA':'Moldova  ( Môn-đô-va )',
    'FIN':'Finland  ( Phần Lan )','FRA':'France  ( Pháp )',
    'DNK':'Denmark  ( Đan Mạch )','MUS':'Mauritius  ( Mô-ri-xơ )',
}
LOAI_GIAY = {
    'Căn cước công dân':'8 - Thẻ Căn Cước','Hộ chiếu':'4 - Hộ chiếu',
    'Chứng minh nhân dân':'2 - Thẻ CMND','Căn cước':'1 - Thẻ CCCD',
    'Giấy khai sinh':'5 - Giấy khai sinh',
}
TINH = {
    'DAK LAK':'605 - Đắk Lắk','DAK NONG':'607 - Đắk Nông','PHU YEN':'509 - Phú Yên',
    'HO CHI MINH':'701 - TP. Hồ Chí Minh','THP HO CHI MINH':'701 - TP. Hồ Chí Minh',
    'HCM':'701 - TP. Hồ Chí Minh','TP HCM':'701 - TP. Hồ Chí Minh',
    'GIA LAI':'603 - Gia Lai','QUANG NGAI':'505 - Quảng Ngãi','LAM DONG':'703 - Lâm Đồng',
    'LAM DONG.':'703 - Lâm Đồng','KHANH HOA':'511 - Khánh Hòa','BEN TRE':'817 - Bến Tre',
    'VINH LONG':'809 - Vĩnh Long','NINH THUAN':'513 - Ninh Thuận','DONG NAI':'713 - Đồng Nai',
    'TIEN GIANG':'807 - Tiền Giang','LONG AN':'801 - Long An','BINH DUONG':'707 - Bình Dương',
    'BINH THUAN':'705 - Bình Thuận','CAN THO':'815 - TP. Cần Thơ','DA NANG':'501 - TP. Đà Nẵng',
    'HA NOI':'101 - TP. Hà Nội','BA RIA VUNG TAU':'711 - Bà Rịa - Vũng Tàu',
    'DONG THAP':'803 - Đồng Tháp','AN GIANG':'805 - An Giang','KIEN GIANG':'819 - Kiên Giang',
    'HA TINH':'407 - Hà Tĩnh','BINH DINH':'507 - Bình Định','VIET NAM':'',
}

# ── Helpers ───────────────────────────────────────────────────────────────
def fmt(v):
    if v is None or str(v) in ('NaT','nan',''): return ''
    if hasattr(v,'strftime'): return v.strftime('%d/%m/%Y')
    return str(v).strip()[:10]

def make_code(prefix, ns):
    if not ns: return prefix
    p = ns.replace('-','/').split('/')
    return f"{prefix}{p[0].zfill(2)}{p[1].zfill(2)}{p[2][-2:]}" if len(p)==3 else prefix

def cp(src, dst):
    for a in ('font','fill','border','alignment'):
        v = getattr(src,a)
        if v: setattr(dst,a,copy(v))
    dst.number_format = src.number_format

def serial2date(s):
    if not s: return None
    try: return datetime.datetime(1899,12,30)+datetime.timedelta(days=int(s))
    except: return None

def wb_to_bytes(wb):
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Processing ────────────────────────────────────────────────────────────
def process_xlsx(xlsx_bytes, rate):
    """Điền dữ liệu file đầu vào lên FILE MẪU customer (QLLT) — giữ nguyên 100%
    template: sheet 'customer' + sheet 'Danh-muc', định dạng header, style ô dữ liệu.
    Map cột theo tên (không phân biệt hoa thường / dấu / khoảng trắng), quy đổi
    ĐƠN GIÁ (USD → VND, tô vàng ô đã đổi), đánh lại STT."""
    src_wb = load_workbook(io.BytesIO(xlsx_bytes))
    src_ws = src_wb.active

    # Map header nguồn (chuẩn hóa) → chỉ số cột nguồn
    src_map = {}
    for c in src_ws[1]:
        if c.value is not None and str(c.value).strip():
            src_map.setdefault(_norm_nat(c.value), c.column)

    # Nạp template QLLT (customer + Danh-muc), dòng 2 là mẫu định dạng ô dữ liệu
    wb = load_workbook(io.BytesIO(load_template('customer')))
    ws = wb['customer']
    n_cols = ws.max_column
    ref = [ws.cell(2, ci) for ci in range(1, n_cols + 1)]
    ref_styles = [(copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment), c.number_format) for c in ref]

    # Với mỗi cột template, tìm cột nguồn tương ứng theo tên
    headers = [ws.cell(1, ci).value for ci in range(1, n_cols + 1)]
    col_src = [src_map.get(_norm_nat(h)) if h else None for h in headers]
    don_gia_idx = next((i + 1 for i, h in enumerate(headers) if h and _norm_nat(h) == _norm_nat('ĐƠN GIÁ')), None)

    ws.delete_rows(2)  # bỏ dòng mẫu

    conv = 0
    er = 1
    for row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row):
        if all(c.value is None or str(c.value).strip() == '' for c in row):
            continue
        er += 1
        for ci in range(1, n_cols + 1):
            cell = ws.cell(er, ci)
            f, fl, b, a, nf = ref_styles[ci - 1]
            cell.font = copy(f); cell.fill = copy(fl); cell.border = copy(b)
            cell.alignment = copy(a); cell.number_format = nf
            sc = col_src[ci - 1]
            if sc is not None and row[sc - 1].value is not None:
                cell.value = row[sc - 1].value
        ws.cell(er, 1).value = er - 1  # STT đánh lại
        if don_gia_idx:
            dg = ws.cell(er, don_gia_idx)
            if dg.value and isinstance(dg.value, (int, float)) and 0 < dg.value < 1000:
                dg.value = round(dg.value * rate)
                dg.fill = PatternFill("solid", start_color="FFFF00")
                conv += 1
    return wb, conv

def split_wb(wb, loai):
    wb2 = load_workbook(io.BytesIO(wb_to_bytes(wb)))
    ws2 = wb2.active
    lc = next(c.column for c in ws2[1] if c.value=='LOẠI KHÁCH')
    dels = [row[0].row for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row)
            if row[lc-1].value != loai]
    for r in reversed(dels): ws2.delete_rows(r)
    for i, row in enumerate(ws2.iter_rows(min_row=2,max_row=ws2.max_row),1): row[0].value=i
    return wb2

def build_kbtt(df_intl):
    wb = load_workbook(io.BytesIO(load_template('kbtt')))
    ws = wb['KBTT']
    ref = [ws.cell(4,c) for c in range(1,12)]
    for r in range(ws.max_row,3,-1): ws.delete_rows(r)
    for i,(_,row) in enumerate(df_intl.iterrows(),1):
        er=i+3
        ht=str(row.get('HỌ TÊN ',row.get('HỌ TÊN',''))).strip()
        ns=fmt(row['NGÀY SINH']); nd=fmt(row['NGÀY ĐẾN']); ni=fmt(row.get('NGÀY ÐI',row.get('NGÀY ĐI','')))
        gt='M - Nam' if str(row.get('GIỚI TÍNH','')).strip()=='Nam' else 'F - Nữ'
        qt=lookup_nat_kbtt(row.get('QUỐC TỊCH',''))
        sh=str(row.get('SỐ GIẤY TỜ','')).strip(); sp=str(row.get('SỐ PHÒNG','')).strip()
        vals=[i,ht,ns,'D - Ngày',gt,qt,sh,sp,nd,ni,ni]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(er,ci); cell.value=val if isinstance(val,int) else str(val)
            cp(ref[ci-1],cell)
    return wb

def build_vnm(df_vn):
    wb = load_workbook(io.BytesIO(load_template('vnm')))
    wsn = next((s for s in wb.sheetnames if 'KHACH' in s or 'DS' in s), wb.sheetnames[0])
    ws = wb[wsn]
    ref = [ws.cell(5,c) for c in range(1,ws.max_column+1)]
    for r in range(ws.max_row,4,-1): ws.delete_rows(r)
    gks_cnt=0; gbl_cnt=0
    for i,(_,row) in enumerate(df_vn.iterrows(),1):
        er=i+4
        ht=str(row.get('HỌ TÊN ',row.get('HỌ TÊN',''))).strip()
        ns=fmt(row['NGÀY SINH']); nd=fmt(row['NGÀY ĐẾN']); ni=fmt(row.get('NGÀY ÐI',row.get('NGÀY ĐI','')))
        gt='F - Nữ' if str(row.get('GIỚI TÍNH','')).strip()=='Nữ' else 'M - Nam'
        sg_raw=str(row.get('SỐ GIẤY TỜ','')).strip()
        lg_raw=str(row.get('LOẠI GIẤY TỜ','')).strip()
        is_gks='GKS' in sg_raw.upper(); is_gbl='GBL' in sg_raw.upper()
        ten_giay=''
        if is_gks:
            sg=make_code('GKS',ns); lg='5 - Giấy khai sinh'; gks_cnt+=1
        elif is_gbl:
            sg=make_code('GBL',ns); lg='9 - Giấy tờ khác'; ten_giay='Giấy bảo lãnh'; gbl_cnt+=1
        elif sg_raw and sg_raw[0].isalpha():
            # Số giấy tờ bắt đầu bằng chữ cái (vd: P02628567) → là hộ chiếu
            sg=sg_raw; lg='4 - Hộ chiếu'
        else:
            sg=sg_raw; lg=LOAI_GIAY.get(lg_raw,lg_raw)
        tinh=TINH.get(str(row.get('TP/TỈNH','')).strip().upper(),'')
        dc=str(row.get('ÐỊA CHỈ',row.get('ĐỊA CHỈ',''))).strip()
        sp=str(row.get('SỐ PHÒNG','')).strip()
        vals=[i,ht,ns,gt,'VNM - Viet Nam',lg,ten_giay,sg,'','1 - Thường trú',tinh,'',dc,nd,ni,sp,'1 - Du lịch','','']
        for ci,val in enumerate(vals,1):
            cell=ws.cell(er,ci); cell.value=val if isinstance(val,int) else str(val)
            if ci<=len(ref): cp(ref[ci-1],cell)
    return wb, gks_cnt, gbl_cnt

def build_dk14(xls_bytes):
    wb2=xlrd.open_workbook(file_contents=xls_bytes)
    ws2=wb2.sheet_by_index(0)
    data=[[ws2.cell_value(r,c) for c in range(ws2.ncols)]
          for r in range(1,ws2.nrows) if any(ws2.cell_value(r,c) for c in range(ws2.ncols))]
    wb_t=load_workbook(io.BytesIO(load_template('dk14')))
    ws_t=wb_t.active
    cw={col:ws_t.column_dimensions[col].width for col in ws_t.column_dimensions}
    rh={r:ws_t.row_dimensions[r].height for r in ws_t.row_dimensions if r<=17}
    wb_o=Workbook(); ws_o=wb_o.active
    for col,w in cw.items(): ws_o.column_dimensions[col].width=w
    for r,h in rh.items():
        if h: ws_o.row_dimensions[r].height=h
    def cc(s,d):
        d.value=s.value
        for a in ('font','fill','border','alignment'):
            v=getattr(s,a)
            if v: setattr(d,a,copy(v))
        d.number_format=s.number_format
    for r in range(1,18):
        for c in range(1,14): cc(ws_t.cell(r,c),ws_o.cell(r,c))
    for mc in ws_t.merged_cells.ranges:
        if mc.min_row<=17:
            ws_o.merge_cells(start_row=mc.min_row,start_column=mc.min_col,
                             end_row=mc.max_row,end_column=mc.max_col)
    wb_t.close()
    thin=Side(style='thin'); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    fn=Font(name='Times New Roman',size=12)
    ac=Alignment(horizontal='center',vertical='center',wrap_text=True)
    al=Alignment(horizontal='left',vertical='center',wrap_text=True)
    for i,row in enumerate(data,1):
        er=i+17
        name=str(row[1]).strip() if row[1] else ''
        gender=str(row[3]).strip().upper() if row[3] else ''
        country=NAT_DK14.get(str(row[4]).strip().upper() if row[4] else '',str(row[4] or ''))
        passport=str(row[5]).strip() if row[5] else ''
        if passport.endswith('.0'): passport=passport[:-2]
        address=str(row[6]).strip() if row[6] else '   '
        dob=serial2date(row[2]); arr=serial2date(row[7]); dep=serial2date(row[8])
        room=str(row[9]).strip() if row[9] else ''
        notify=str(row[10]).strip() if row[10] else ''
        cols=[(1,i,ac),(2,name,al),(3,dob if gender=='M' else None,ac),
              (4,dob if gender=='F' else None,ac),(5,country,ac),(6,passport,ac),
              (7,address,al),(8,arr,ac),(9,dep,ac),(10,room,ac),(11,notify,al),(12,'',ac),(13,'',ac)]
        for ci,val,aln in cols:
            cell=ws_o.cell(er,ci); cell.value=val; cell.font=fn; cell.border=bdr; cell.alignment=aln
            if ci in (3,4) and val: cell.number_format='DD/MM/YYYY'
            elif ci in (8,9) and val: cell.number_format='DD/MM/YYYY'
    return wb_o, len(data)


# ── Regcard PDF builder ───────────────────────────────────────────────────
@st.cache_resource
def load_regcard_template():
    path = os.path.join(os.path.dirname(__file__), 'tmpl_regcard.b64')
    with open(path, 'r') as f:
        return base64.b64decode(f.read())

def _rc_clean_name(n):
    if pd.isna(n): return ''
    return str(n).strip().rstrip(',').strip()

def _rc_conf(c):
    if pd.isna(c): return ''
    return str(int(c)) if isinstance(c,(int,float)) else str(c)

def _rc_date(d):
    if pd.isna(d): return ''
    if hasattr(d,'strftime'):
        return f"{d.month:02d}/{d.day:02d}/{d.year}"
    s=str(d).strip(); p=s.split('/')
    if len(p)==3:
        dd,mm,yy=p
        if len(yy)==2: yy='20'+yy
        return f"{dd}/{mm}/{yy}"
    return s

def _rc_nights(arr,dep):
    try:
        if hasattr(arr,'strftime'):
            a=pd.Timestamp(year=arr.year,month=arr.day,day=arr.month)
        else:
            p=str(arr).split('/'); a=pd.Timestamp(f"20{p[2]}-{p[1]}-{p[0]}")
        p=str(dep).split('/')
        d=pd.Timestamp(f"20{p[2]}-{p[1]}-{p[0]}") if len(p[2])==2 else pd.to_datetime(dep)
        return str((d-a).days)
    except: return ''

def build_regcards(xlsx_bytes, only_main=True):
    """Tạo PDF regcard hàng loạt, gộp theo Conf# (đoàn nhiều phòng → 1 regcard,
    các số phòng gộp chung vào ô Room No)."""
    df = pd.read_excel(io.BytesIO(xlsx_bytes))
    H = 841.0
    FONT = "Times-Roman"; SIZE = 9.8
    # Baseline chính xác (bottom) đo từ dữ liệu mẫu gốc — chữ trùng khít 100%
    POS = {
        'name':(125.50,109.60),'conf':(526.75,108.52),'arrival':(119.90,144.22),
        'departure':(329.90,144.22),'nights':(500.30,144.22),'type':(113.60,179.92),
        'rm':(360.60,179.92),'company':(221.40,216.32),
    }
    # Ô che dữ liệu cũ — vừa khít vùng chữ, không lấn đường kẻ bảng
    BLANK = [
        (125,99,250,110.5),(526,98,568,109.5),(119,133.5,168,145.2),
        (329,133.5,378,145.2),(500,133.5,512,145.2),(113,169,145,180.9),
        (360,169,410,180.9),(221,205.5,320,217.3),
    ]

    # ── Gộp theo Conf# ──
    # Điền mã Conf# xuống các dòng trống (khách đi cùng booking),
    # rồi gộp TẤT CẢ dòng cùng 1 mã Conf# thành 1 regcard, gộp mọi số phòng.
    def _clean_room(r):
        s = str(r).strip()
        if s.endswith('.0'): s = s[:-2]
        return s

    df = df.copy()
    df['_conf_ff'] = df['Conf#'].ffill()
    groups = []
    for conf_val, grp in df.groupby('_conf_ff', sort=False):
        main_rows = grp[grp['Conf#'].notna()]
        if len(main_rows) == 0:
            continue
        main = main_rows.iloc[0]
        rooms = []
        for r in grp['Rm']:
            if pd.notna(r):
                rs = _clean_room(r)
                # Bỏ phòng ảo đầu 9 dạng 9000-9999 (9002, 9005, 9010, 9040... — posting master)
                if rs and rs not in rooms and not _re.fullmatch(r'9\d{3}', rs):
                    rooms.append(rs)
        groups.append({'main': main, 'rooms': rooms})

    tmpl_bytes = load_regcard_template()
    writer = PdfWriter()
    count = 0
    for g in groups:
        row = g['main']
        name = _rc_clean_name(row.get('Name'))
        if not name:
            continue
        data = {
            'name': name,
            'conf': _rc_conf(row.get('Conf#')),
            'arrival': _rc_date(row.get('Arrival')),
            'departure': _rc_date(row.get('Departure')),
            'nights': _rc_nights(row.get('Arrival'), row.get('Departure')),
            'type': str(row.get('Type')) if pd.notna(row.get('Type')) else '',
            'rm': ', '.join(g['rooms']),   # gộp các số phòng
            'company': str(row.get('Company')) if pd.notna(row.get('Company')) else '',
        }
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=(595,841))
        c.setFillColor(white)
        for x0,top,x1,bot in BLANK:
            c.rect(x0, H-bot, (x1-x0), (bot-top), fill=1, stroke=0)
        c.setFillColor(black)
        c.setFont(FONT, SIZE)
        MAXW = {'name': 300, 'company': 200}  # ô 1 dòng: thu nhỏ nếu tràn
        for key,(x,bottom) in POS.items():
            val = data[key]
            if not val:
                continue
            if key == 'rm':
                # Ô số phòng: 1 phòng → vị trí chuẩn như mẫu gốc.
                # Nhiều phòng → căn giữa CHÍNH XÁC giữa 2 nhãn (đo từ 2 ảnh crop độc lập,
                # sai lệch <1pt): nhãn "ROOM NO./(Số phòng)" kết thúc ~261pt,
                # nhãn "ROOM RATE/(Giá phòng)" bắt đầu ~441pt.
                # → Vùng vẽ 266–436pt, TÂM = 351pt, rộng 170pt
                #   (đủ ~8 phòng/dòng ở cỡ chữ nguyên vẹn 9.8pt → 24 phòng/3 dòng).
                rooms = [s.strip() for s in val.split(',') if s.strip()]
                RM_X0, RM_X1 = 266.0, 436.0
                RM_CX = (RM_X0 + RM_X1) / 2        # = 351.0
                RM_MAXW = RM_X1 - RM_X0            # = 170pt mỗi dòng
                def _wrap(items, fs):
                    lines=[]; cur=''
                    for it in items:
                        test = (cur + ', ' + it) if cur else it
                        if c.stringWidth(test, FONT, fs) <= RM_MAXW:
                            cur = test
                        else:
                            if cur: lines.append(cur)
                            cur = it
                    if cur: lines.append(cur)
                    return lines
                if len(rooms) <= 1:
                    c.drawString(x, H - bottom, val)
                else:
                    fs = SIZE
                    lines = _wrap(rooms, fs)
                    if len(lines) <= 3:
                        gap = fs + 1.6             # tới ~24 phòng: cỡ chữ 9.8pt nguyên vẹn
                    else:
                        gap = fs + 0.6             # 4 dòng sát nhau, vẫn nguyên cỡ chữ (~32 phòng)
                        while len(lines) > 4 and fs > 7:
                            fs -= 0.3              # chống tràn tuyệt đối cho case phi thực tế
                            lines = _wrap(rooms, fs)
                    c.setFont(FONT, fs)
                    for i, ln in enumerate(lines):
                        c.drawCentredString(RM_CX, H - bottom - i*gap, ln)
                    c.setFont(FONT, SIZE)
                    c.setFont(FONT, fs)
                    for i, ln in enumerate(lines):
                        c.drawCentredString(RM_CX, H - bottom - i*gap, ln)
                    c.setFont(FONT, SIZE)
            else:
                maxw = MAXW.get(key)
                if maxw:
                    fs = SIZE
                    while fs > 6 and c.stringWidth(val, FONT, fs) > maxw:
                        fs -= 0.3
                    c.setFont(FONT, fs)
                    c.drawString(x, H-bottom, val)
                    c.setFont(FONT, SIZE)
                else:
                    c.drawString(x, H-bottom, val)
        c.save(); buf.seek(0)
        base = PdfReader(io.BytesIO(tmpl_bytes))
        overlay = PdfReader(buf)
        page = base.pages[0]
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
        count += 1
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue(), count

# ── Đối chiếu Smile vs Trang lưu trú người nước ngoài ─────────────────────
def _norm_pp(p):
    """Chuẩn hóa số hộ chiếu."""
    if pd.isna(p): return ''
    s = str(p).strip().upper().replace(' ','')
    if s.endswith('.0'): s = s[:-2]
    return s

def _norm_room(r):
    if pd.isna(r): return ''
    s = str(r).strip()
    if s.endswith('.0'): s = s[:-2]
    return s.upper()  # 12a05 và 12A05 là một phòng

def reconcile(smile_bytes, luutru_bytes, today):
    """Đối chiếu file Smile (inhouse) với file trang quản lý lưu trú.
    today: pd.Timestamp ngày xuất file (hôm nay)."""
    # ── Đọc Smile ──
    df1 = pd.read_excel(io.BytesIO(smile_bytes), header=0)
    smile = df1[['Passport #','NAT','Rm#','Arrival','Last Name','First Name']].copy()
    smile = smile.dropna(subset=['Passport #'])
    smile['pp'] = smile['Passport #'].apply(_norm_pp)
    smile['Arrival'] = pd.to_datetime(smile['Arrival'], errors='coerce')
    smile['name'] = (smile['Last Name'].astype(str).str.strip() + ' ' +
                     smile['First Name'].astype(str).str.strip())
    smile['room'] = smile['Rm#'].apply(_norm_room)
    # Cột Departure (dò tên linh hoạt phòng khi khác tên)
    _dep_col = next((c for c in df1.columns if 'depart' in str(c).lower()), None)
    smile['Departure'] = pd.to_datetime(df1[_dep_col], errors='coerce') if _dep_col else pd.NaT
    # Lọc: bỏ VNM + bỏ arrival = hôm nay + bỏ departure = hôm nay (khách trả phòng)
    smile_f = smile[(smile['NAT'] != 'VNM') &
                    (smile['Arrival'].dt.date != today.date()) &
                    (smile['Departure'].dt.date != today.date())].copy()

    # ── Đọc Lưu trú ──
    df2 = pd.read_excel(io.BytesIO(luutru_bytes), header=9)
    df2 = df2.dropna(subset=['Họ tên'])
    df2['pp'] = df2['Số hộ chiếu'].apply(_norm_pp)
    df2['ddk'] = pd.to_datetime(df2['Ngày đi dự kiến'], format='%d/%m/%Y', errors='coerce')
    df2['room'] = df2['Số phòng'].apply(_norm_room)
    # Lọc: bỏ ngày đi dự kiến = hôm nay
    luutru_f = df2[df2['ddk'].dt.date != today.date()].copy()

    smile_pp = set(smile_f['pp'])
    luutru_pp = set(luutru_f['pp'])

    # Đối chiếu người
    chua_dk = smile_f[~smile_f['pp'].isin(luutru_pp)][['name','pp','NAT','room','Arrival']].copy()
    chua_dk.columns = ['Họ tên','Số hộ chiếu','Quốc tịch','Số phòng','Ngày đến']
    chua_dk['Ngày đến'] = chua_dk['Ngày đến'].dt.strftime('%d/%m/%Y')

    thua = luutru_f[~luutru_f['pp'].isin(smile_pp)].copy()
    # Giai đoạn lưu trú: Ngày đến → Ngày đi dự kiến
    _dd = pd.to_datetime(thua['Ngày đến '], format='%d/%m/%Y', errors='coerce') if 'Ngày đến ' in thua else pd.Series([pd.NaT]*len(thua))
    thua['_luutru'] = (_dd.dt.strftime('%d/%m/%Y').fillna('?') + ' → ' +
                       thua['ddk'].dt.strftime('%d/%m/%Y').fillna('?'))
    _qt = thua['QT'] if 'QT' in thua else ''
    thua = pd.DataFrame({
        'Họ tên': thua['Họ tên'].values,
        'Số hộ chiếu': thua['pp'].values,
        'Quốc tịch': _qt.values if hasattr(_qt,'values') else _qt,
        'Số phòng': thua['room'].values,
        'Giai đoạn lưu trú': thua['_luutru'].values,
    })

    dup = luutru_f[luutru_f['pp'].duplicated(keep=False) & (luutru_f['pp']!='')]
    dup = dup[['Họ tên','pp','room']].sort_values('pp').copy()
    dup.columns = ['Họ tên','Số hộ chiếu','Số phòng']

    # Đối chiếu phòng
    smile_rooms = set(r for r in smile_f['room'] if r)
    luutru_rooms = set(r for r in luutru_f['room'] if r)
    def _sortkey(x): return (len(x), x)
    room_chua = sorted(smile_rooms - luutru_rooms, key=_sortkey)
    room_thua = sorted(luutru_rooms - smile_rooms, key=_sortkey)

    return {
        'smile_total': len(smile), 'smile_filtered': len(smile_f),
        'luutru_total': len(df2), 'luutru_filtered': len(luutru_f),
        'chua_dk': chua_dk, 'thua': thua, 'dup': dup,
        'room_chua': room_chua, 'room_thua': room_thua,
        'room_match': len(smile_rooms & luutru_rooms),
        'smile_rooms': len(smile_rooms), 'luutru_rooms': len(luutru_rooms),
    }



# ── UI ────────────────────────────────────────────────────────────────────

# Custom CSS — Minimalist white (Notion/Linear style)
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header[data-testid="stHeader"] {background: transparent;}
    .stApp {
        background: url("data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAxNDQwIDkwMCIgcHJlc2VydmVBc3BlY3RSYXRpbz0ieE1pZFlNaWQgc2xpY2UiPjxkZWZzPjxsaW5lYXJHcmFkaWVudCBpZD0ic2t5IiB4MT0iMCIgeTE9IjAiIHgyPSIwIiB5Mj0iMSI+PHN0b3Agb2Zmc2V0PSIwJSIgc3RvcC1jb2xvcj0iIzlmZDRmNSIvPjxzdG9wIG9mZnNldD0iNTAlIiBzdG9wLWNvbG9yPSIjYzllOGZhIi8+PHN0b3Agb2Zmc2V0PSIxMDAlIiBzdG9wLWNvbG9yPSIjZWFmNmZkIi8+PC9saW5lYXJHcmFkaWVudD48bGluZWFyR3JhZGllbnQgaWQ9Im10biIgeDE9IjAiIHkxPSIwIiB4Mj0iMCIgeTI9IjEiPjxzdG9wIG9mZnNldD0iMCUiIHN0b3AtY29sb3I9IiM3ZjljYzAiLz48c3RvcCBvZmZzZXQ9IjEwMCUiIHN0b3AtY29sb3I9IiM1ZjdmYTgiLz48L2xpbmVhckdyYWRpZW50PjxsaW5lYXJHcmFkaWVudCBpZD0ibXRuMiIgeDE9IjAiIHkxPSIwIiB4Mj0iMCIgeTI9IjEiPjxzdG9wIG9mZnNldD0iMCUiIHN0b3AtY29sb3I9IiNiOGNmZTYiLz48c3RvcCBvZmZzZXQ9IjEwMCUiIHN0b3AtY29sb3I9IiM5ZmJkZDkiLz48L2xpbmVhckdyYWRpZW50PjxsaW5lYXJHcmFkaWVudCBpZD0id2F0ZXIiIHgxPSIwIiB5MT0iMCIgeDI9IjAiIHkyPSIxIj48c3RvcCBvZmZzZXQ9IjAlIiBzdG9wLWNvbG9yPSIjYmZlMGY1Ii8+PHN0b3Agb2Zmc2V0PSIxMDAlIiBzdG9wLWNvbG9yPSIjN2ZiNGRkIi8+PC9saW5lYXJHcmFkaWVudD48L2RlZnM+PHJlY3Qgd2lkdGg9IjE0NDAiIGhlaWdodD0iOTAwIiBmaWxsPSJ1cmwoI3NreSkiLz48Y2lyY2xlIGN4PSIxMDgwIiBjeT0iMTYwIiByPSIxMjAiIGZpbGw9IiNmZmYzYjAiIG9wYWNpdHk9IjAuMTgiLz48Y2lyY2xlIGN4PSIxMDgwIiBjeT0iMTYwIiByPSI4MCIgZmlsbD0iI2ZmZjNiMCIgb3BhY2l0eT0iMC45Ii8+PHBhdGggZD0iTTAgNTYwIFEgMjQwIDUwMCA0ODAgNTQ1IFQgOTYwIDUzMCBUIDE0NDAgNTU1IEwxNDQwIDkwMCBMMCA5MDAgWiIgZmlsbD0idXJsKCNtdG4yKSIgb3BhY2l0eT0iMC41NSIvPjxwYXRoIGQ9Ik0tNDAgNjMwIEwzMDAgMjUwIEw2NDAgNjMwIFoiIGZpbGw9InVybCgjbXRuKSIvPjxwYXRoIGQ9Ik0yMTYgMzYxIEwzMDAgMjUwIEwzOTEgMzY3IFEgMzUyIDM0NCAzMjYgMzcwIFEgMzAwIDM0NCAyNzQgMzcwIFEgMjQ4IDM0OCAyMTYgMzYxIFoiIGZpbGw9IiNmYmZkZmYiLz48cGF0aCBkPSJNMjYxIDQxOSBMMjgxIDQwNiBMMjcxIDQ1MiBMMjU1IDQ1OCBaIiBmaWxsPSIjZmJmZGZmIiBvcGFjaXR5PSIwLjkiLz48cGF0aCBkPSJNMzMzIDQxNiBMMzIwIDQwMyBMMzM2IDQ0OCBMMzUyIDQ1NSBaIiBmaWxsPSIjZmJmZGZmIiBvcGFjaXR5PSIwLjkiLz48ZWxsaXBzZSBjeD0iMzAwIiBjeT0iNjM4IiByeD0iMzYwIiByeT0iMjYiIGZpbGw9IiNmZmZmZmYiIG9wYWNpdHk9IjAuNDUiLz48cGF0aCBkPSJNMCA2NjAgUSAzNjAgNjMwIDcyMCA2NTUgVCAxNDQwIDY2MCBMMTQ0MCA5MDAgTDAgOTAwIFoiIGZpbGw9InVybCgjd2F0ZXIpIi8+PHBhdGggZD0iTTAgNjYwIFEgMzYwIDYzMCA3MjAgNjU1IFQgMTQ0MCA2NjAiIGZpbGw9Im5vbmUiIHN0cm9rZT0iI2ZmZmZmZiIgc3Ryb2tlLXdpZHRoPSIyIiBvcGFjaXR5PSIwLjM1Ii8+PHBhdGggZD0iTS00MCA2NzIgTDMwMCA4NTIgTDY0MCA2NzIgWiIgZmlsbD0idXJsKCNtdG4pIiBvcGFjaXR5PSIwLjIyIi8+PHBhdGggZD0iTTI0MCA3MjYgTDMwMCA4MDAgTDM2NCA3MzAgUSAzMzYgNzQ1IDMyMiA3MjggUSAzMDAgNzQ2IDI4MiA3MjggUSAyNjIgNzQwIDI0MCA3MjYgWiIgZmlsbD0iI2ZiZmRmZiIgb3BhY2l0eT0iMC4xNCIvPjxlbGxpcHNlIGN4PSIxMDgwIiBjeT0iNzA1IiByeD0iNzIiIHJ5PSIxMyIgZmlsbD0iI2ZmZjNiMCIgb3BhY2l0eT0iMC4zIi8+PGVsbGlwc2UgY3g9IjEwNjgiIGN5PSI3MzUiIHJ4PSI0NiIgcnk9IjgiIGZpbGw9IiNmZmYzYjAiIG9wYWNpdHk9IjAuMTgiLz48ZWxsaXBzZSBjeD0iMTA5MCIgY3k9Ijc2MiIgcng9IjI4IiByeT0iNSIgZmlsbD0iI2ZmZjNiMCIgb3BhY2l0eT0iMC4xIi8+PHBhdGggZD0iTTEyMCA3MzAgcSA0NSAtOCA5MCAwIiBmaWxsPSJub25lIiBzdHJva2U9IiNmZmZmZmYiIHN0cm9rZS13aWR0aD0iMi41IiBvcGFjaXR5PSIwLjE0IiBzdHJva2UtbGluZWNhcD0icm91bmQiLz48cGF0aCBkPSJNNTQwIDc4MCBxIDU1IC05IDExMCAwIiBmaWxsPSJub25lIiBzdHJva2U9IiNmZmZmZmYiIHN0cm9rZS13aWR0aD0iMi41IiBvcGFjaXR5PSIwLjEyIiBzdHJva2UtbGluZWNhcD0icm91bmQiLz48cGF0aCBkPSJNODIwIDcyMCBxIDQwIC03IDgwIDAiIGZpbGw9Im5vbmUiIHN0cm9rZT0iI2ZmZmZmZiIgc3Ryb2tlLXdpZHRoPSIyLjUiIG9wYWNpdHk9IjAuMTQiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIvPjxwYXRoIGQ9Ik0zMDAgODMwIHEgNjAgLTkgMTIwIDAiIGZpbGw9Im5vbmUiIHN0cm9rZT0iI2ZmZmZmZiIgc3Ryb2tlLXdpZHRoPSIyLjUiIG9wYWNpdHk9IjAuMSIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIi8+PHBhdGggZD0iTTExNTAgODAwIHEgNTAgLTggMTAwIDAiIGZpbGw9Im5vbmUiIHN0cm9rZT0iI2ZmZmZmZiIgc3Ryb2tlLXdpZHRoPSIyLjUiIG9wYWNpdHk9IjAuMTIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIvPjxnIG9wYWNpdHk9IjAuNSIgZmlsbD0iI2YyYzlkOCI+PGNpcmNsZSBjeD0iMTgwIiBjeT0iNzAwIiByPSI3Ii8+PGNpcmNsZSBjeD0iMjA1IiBjeT0iNjkwIiByPSI2Ii8+PGNpcmNsZSBjeD0iMTk1IiBjeT0iNzEyIiByPSI2Ii8+PGNpcmNsZSBjeD0iMTI2MCIgY3k9IjY4MCIgcj0iNyIvPjxjaXJjbGUgY3g9IjEyODUiIGN5PSI2OTUiIHI9IjYiLz48Y2lyY2xlIGN4PSIxMjcwIiBjeT0iNjY4IiByPSI1Ii8+PC9nPjwvc3ZnPg==") center 35% / cover no-repeat fixed, #eaf4fb;
    }
    /* Khối nội dung: kính mờ (frosted glass) — đủ đục để chữ luôn dễ đọc trên MỌI nền
       (sáng/trưa/tối), cảnh Phú Sĩ + trăng nằm hai bên ngoài khung nên vẫn thấy trọn */
    .block-container {
        background: rgba(255,255,255,0.55);
        backdrop-filter: blur(14px) saturate(1.05);
        -webkit-backdrop-filter: blur(14px) saturate(1.05);
        border: 1px solid rgba(255,255,255,0.65);
        border-radius: 18px;
        box-shadow: 0 8px 32px rgba(31,52,96,0.14);
        margin-top: 1.5rem; margin-bottom: 1.5rem;
        position: relative; z-index: 5;
    }
    .block-container {padding-top: 2.5rem; padding-bottom: 2.5rem; max-width: 720px;}
    /* Đưa khung công việc ra GIỮA màn hình theo chiều dọc */
    section[data-testid="stMain"] > div,
    div[data-testid="stMainBlockContainer"] { }
    section[data-testid="stMain"] {
        display: flex; flex-direction: column; justify-content: center;
        min-height: 100vh;
    }
    .app-header {
        margin-bottom: 2.5rem; padding-bottom: 1.5rem;
        border-bottom: 1px solid #ececec;
    }
    .greet-row {
        display: flex; align-items: center; justify-content: space-between;
        flex-wrap: wrap; gap: 8px;
    }
    .greet-text {
        font-size: 1.5rem; font-weight: 650; color: #1a1a1a;
        letter-spacing: -0.02em; margin: 0;
        animation: fadeIn 0.5s ease both;
    }
    .greet-emoji {margin-right: 6px;}
    .clock-text {
        font-size: 0.85rem; color: #9b9b9b; font-weight: 500;
        font-variant-numeric: tabular-nums;
        animation: fadeIn 0.5s ease 0.1s both;
    }
    #live-clock {color: #6b6b6b; font-weight: 600;}
    @keyframes fadeIn {from {opacity: 0; transform: translateY(6px);} to {opacity: 1; transform: translateY(0);}}
    .section-label {
        font-size: 0.7rem; font-weight: 650; color: #6b6b7e;
        text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 0.75rem;
    }
    .menu-card {
        border-radius: 12px; padding: 1.5rem 1.4rem; margin-bottom: 0.6rem;
        border: 1px solid rgba(255,255,255,0.7);
        background: rgba(255,255,255,0.85);
        backdrop-filter: blur(6px); -webkit-backdrop-filter: blur(6px);
        transition: all 0.18s ease; min-height: 150px;
        animation: fadeIn 0.5s ease 0.15s both;
    }
    .menu-card:hover {
        background: rgba(255,255,255,0.95);
        border-color: rgba(255,255,255,0.8); box-shadow: 0 2px 14px rgba(31,52,96,0.10);
    }
    .menu-icon {font-size: 1.9rem; margin-bottom: 0.7rem; display: block;}
    .menu-title {font-size: 1.02rem; font-weight: 650; color: #1a1a1a; margin-bottom: 0.35rem; letter-spacing: -0.01em;}
    .menu-desc {font-size: 0.82rem; color: #8b8b8b; line-height: 1.55;}
    div[data-testid="stFileUploader"] {
        border: 1px dashed #dcdcdc; border-radius: 10px;
        padding: 0.4rem; background: #fcfcfc;
    }
    .stTextInput input, .stNumberInput input {
        border-radius: 8px !important; border-color: #e4e4e4 !important;
    }
    .stButton button {
        border-radius: 8px; font-weight: 550; padding: 0.6rem;
        border: 1px solid #e4e4e4; transition: all 0.15s ease;
    }
    .stButton button:hover {border-color: #b4b4b4; background: #fafafa;}
    .stButton button[kind="primary"] {
        background: #1a1a1a; border-color: #1a1a1a; color: #fff;
    }
    .stButton button[kind="primary"]:hover, .stFormSubmitButton button:hover {background: #333; border-color: #333;}
    .stDownloadButton button {
        border-radius: 8px; font-weight: 600; padding: 0.75rem;
        background: #1a1a1a; border-color: #1a1a1a; color: #fff;
    }
    .stDownloadButton button:hover {background: #333;}
    div[data-testid="stMetric"] {
        background: #ffffff; border: 1px solid #ececec;
        border-radius: 10px; padding: 0.85rem 0.5rem; text-align: center;
    }
    div[data-testid="stMetricValue"] {color: #1a1a1a; font-weight: 700; font-size: 1.4rem;}
    div[data-testid="stMetricLabel"] {color: #9b9b9b;}
    .login-box {text-align: center; padding: 2rem 0 1rem;}
    .login-box .lb-logo {font-size: 2.4rem;}
    .login-box h2 {color: #1a1a1a; font-size: 1.3rem; font-weight: 700; margin: 0.6rem 0 0.25rem; letter-spacing: -0.02em;}
    .login-box p {color: #9b9b9b; font-size: 0.85rem; margin: 0;}
</style>
""", unsafe_allow_html=True)

# Minimalist welcome header — lời chào + đồng hồ tính theo GIỜ MÁY NGƯỜI DÙNG (JS)
# Splash chào mừng hoạt hình — chỉ hiện 1 lần mỗi phiên tab (sessionStorage), click để bỏ qua
components.html("""
<script>
(function(){
    var doc = window.parent.document;
    var win = window.parent;
    try {
        if (win.sessionStorage.getItem('hp_splash_shown')) return;
        win.sessionStorage.setItem('hp_splash_shown', '1');
    } catch(e) { return; }

    var old = doc.getElementById('splash-overlay'); if (old) old.remove();
    var oldCss = doc.getElementById('splash-style'); if (oldCss) oldCss.remove();

    // Danh sách lời chào — chỉ Tiếng Việt và Tiếng Anh (kiểu iPhone)
    var GREETINGS = ["Xin chào", "Hello", "Xin chào"];

    var css = doc.createElement('style');
    css.id = 'splash-style';
    css.textContent = `
      #splash-overlay{position:fixed;inset:0;z-index:999999;overflow:hidden;cursor:pointer;
        display:flex;align-items:center;justify-content:center;
        background:#8a9bb8;
        transition:opacity .8s ease;}
      #splash-overlay.splash-exit{opacity:0;pointer-events:none;}
      /* Nền gradient nhiều màu chuyển động mượt (kiểu iOS) */
      #splash-overlay .mesh{position:absolute;inset:-40%;filter:blur(70px);opacity:.9;
        background:
          radial-gradient(closest-side at 30% 35%, #f0c86e, transparent 70%),
          radial-gradient(closest-side at 70% 30%, #5b8def, transparent 70%),
          radial-gradient(closest-side at 65% 70%, #7fb2c8, transparent 70%),
          radial-gradient(closest-side at 30% 72%, #e9a13b, transparent 70%),
          radial-gradient(closest-side at 50% 50%, #8bb0d0, transparent 75%);
        animation:mesh-move 14s ease-in-out infinite alternate;}
      @keyframes mesh-move{
        0%{transform:translate(0,0) rotate(0deg) scale(1.05);}
        50%{transform:translate(3%,-3%) rotate(8deg) scale(1.18);}
        100%{transform:translate(-3%,3%) rotate(-6deg) scale(1.08);}
      }
      #splash-hello{position:relative;z-index:2;
        font-family:'Segoe Script','Brush Script MT','Snell Roundhand',cursive;
        font-size:clamp(3rem,9vw,5.4rem);font-weight:400;letter-spacing:0.01em;
        color:#ffffff;text-align:center;padding:0 6vw;
        text-shadow:0 2px 18px rgba(0,0,0,.18);white-space:pre;}
      #splash-hello .ch{opacity:0;display:inline-block;
        animation:ch-in .5s ease forwards;}
      @keyframes ch-in{from{opacity:0;transform:translateY(12px);}to{opacity:1;transform:translateY(0);}}
      #splash-hello.fade-out{animation:hello-out .5s ease forwards;}
      @keyframes hello-out{from{opacity:1;}to{opacity:0;transform:translateY(-12px);}}
    `;
    doc.head.appendChild(css);

    var s = doc.createElement('div');
    s.id = 'splash-overlay';
    s.innerHTML = '<div class="mesh"></div><div id="splash-hello"></div>';
    doc.body.appendChild(s);

    var el = s.querySelector('#splash-hello');
    var i = 0, timers = [];
    var STEP = 130;   // thời gian mỗi ký tự hiện ra (ms) — chữ chạy từ từ
    var HOLD = 850;   // giữ nguyên chữ sau khi gõ xong (ms)

    function showNext(){
        if (i >= GREETINGS.length){ dismiss(); return; }
        var word = GREETINGS[i]; i++;
        // Dựng từng ký tự, mỗi ký tự hiện lần lượt cách nhau STEP ms
        el.className = '';
        el.innerHTML = '';
        for (var k = 0; k < word.length; k++){
            var span = doc.createElement('span');
            span.className = 'ch';
            span.textContent = word[k];
            span.style.animationDelay = (k * STEP) + 'ms';
            el.appendChild(span);
        }
        var typeDur = word.length * STEP;
        // Gõ xong → giữ → fade ra → chữ tiếp theo
        timers.push(win.setTimeout(function(){
            el.classList.add('fade-out');
            timers.push(win.setTimeout(showNext, 450));
        }, typeDur + HOLD));
    }
    showNext();

    function dismiss(){
        timers.forEach(function(t){ win.clearTimeout(t); });
        s.classList.add('splash-exit');
        win.setTimeout(function(){ s.remove(); var c = doc.getElementById('splash-style'); if (c) c.remove(); }, 750);
    }
    s.addEventListener('click', dismiss);  // chạm để bỏ qua
    // Chốt an toàn: ép gỡ overlay sau 12s dù có sự cố (tránh chặn thao tác)
    win.setTimeout(function(){
        var sp = doc.getElementById('splash-overlay'); if (sp) sp.remove();
        var c = doc.getElementById('splash-style'); if (c) c.remove();
    }, 12000);
})();
</script>
""", height=0)

# ---- Nền đổi theo thời gian trong ngày (sáng / trưa-chiều / tối) — theo giờ máy người dùng ----
import base64 as _b64

def _fuji_svg(p):
    """Tạo SVG cảnh núi Phú Sĩ với bảng màu p."""
    stars = ""
    if p.get("stars"):
        import random as _rd
        _rd.seed(7)  # cố định vị trí sao để base64 không đổi giữa các rerun
        dots = "".join(
            '<circle cx="%d" cy="%d" r="%.1f" fill="#ffffff" opacity="%.2f"/>'
            % (_rd.randint(30, 1410), _rd.randint(20, 260), _rd.uniform(1.2, 2.6), _rd.uniform(0.4, 0.95))
            for _ in range(26)
        )
        stars = '<g>' + dots + '</g>'
    return (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1440 900" preserveAspectRatio="xMidYMid slice">'
        '<defs>'
        '<linearGradient id="sky" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{p["sky0"]}"/><stop offset="50%" stop-color="{p["sky1"]}"/><stop offset="100%" stop-color="{p["sky2"]}"/>'
        '</linearGradient>'
        '<linearGradient id="mtn" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{p["mtn0"]}"/><stop offset="100%" stop-color="{p["mtn1"]}"/>'
        '</linearGradient>'
        '<linearGradient id="mtn2" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{p["far0"]}"/><stop offset="100%" stop-color="{p["far1"]}"/>'
        '</linearGradient>'
        '<linearGradient id="water" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{p["water0"]}"/><stop offset="100%" stop-color="{p["water1"]}"/>'
        '</linearGradient>'
        '</defs>'
        '<rect width="1440" height="900" fill="url(#sky)"/>'
        + stars +
        f'<circle cx="{p["orb_x"]}" cy="{p["orb_y"]}" r="{p["orb_r"] + 40}" fill="{p["orb"]}" opacity="0.18"/>'
        f'<circle cx="{p["orb_x"]}" cy="{p["orb_y"]}" r="{p["orb_r"]}" fill="{p["orb"]}" opacity="{p["orb_op"]}"/>'
        '<path d="M0 560 Q 240 500 480 545 T 960 530 T 1440 555 L1440 900 L0 900 Z" fill="url(#mtn2)" opacity="0.55"/>'
        # Núi Phú Sĩ (phóng to, lệch trái)
        '<path d="M-40 630 L300 250 L640 630 Z" fill="url(#mtn)"/>'
        f'<path d="M216 361 L300 250 L391 367 Q 352 344 326 370 Q 300 344 274 370 Q 248 348 216 361 Z" fill="{p["snow"]}"/>'
        f'<path d="M261 419 L281 406 L271 452 L255 458 Z" fill="{p["snow"]}" opacity="0.9"/>'
        f'<path d="M333 416 L320 403 L336 448 L352 455 Z" fill="{p["snow"]}" opacity="0.9"/>'
        f'<ellipse cx="300" cy="638" rx="360" ry="26" fill="#ffffff" opacity="{p["mist_op"]}"/>'
        # Mặt hồ tiền cảnh
        '<path d="M0 660 Q 360 630 720 655 T 1440 660 L1440 900 L0 900 Z" fill="url(#water)"/>'
        '<path d="M0 660 Q 360 630 720 655 T 1440 660" fill="none" stroke="#ffffff" stroke-width="2" opacity="0.35"/>'
        # Bóng núi phản chiếu (lộn ngược, nén dọc, mờ)
        '<path d="M-40 672 L300 852 L640 672 Z" fill="url(#mtn)" opacity="0.22"/>'
        f'<path d="M240 726 L300 800 L364 730 Q 336 745 322 728 Q 300 746 282 728 Q 262 740 240 726 Z" fill="{p["snow"]}" opacity="0.14"/>'
        # Bóng trăng/mặt trời lung linh trên nước
        f'<ellipse cx="{p["orb_x"]}" cy="705" rx="72" ry="13" fill="{p["orb"]}" opacity="0.3"/>'
        f'<ellipse cx="{p["orb_x"] - 12}" cy="735" rx="46" ry="8" fill="{p["orb"]}" opacity="0.18"/>'
        f'<ellipse cx="{p["orb_x"] + 10}" cy="762" rx="28" ry="5" fill="{p["orb"]}" opacity="0.1"/>'
        # Gợn sóng nhẹ
        '<path d="M120 730 q 45 -8 90 0" fill="none" stroke="#ffffff" stroke-width="2.5" opacity="0.14" stroke-linecap="round"/>'
        '<path d="M540 780 q 55 -9 110 0" fill="none" stroke="#ffffff" stroke-width="2.5" opacity="0.12" stroke-linecap="round"/>'
        '<path d="M820 720 q 40 -7 80 0" fill="none" stroke="#ffffff" stroke-width="2.5" opacity="0.14" stroke-linecap="round"/>'
        '<path d="M300 830 q 60 -9 120 0" fill="none" stroke="#ffffff" stroke-width="2.5" opacity="0.1" stroke-linecap="round"/>'
        '<path d="M1150 800 q 50 -8 100 0" fill="none" stroke="#ffffff" stroke-width="2.5" opacity="0.12" stroke-linecap="round"/>'
        # Cánh hoa trôi trên mặt nước
        f'<g opacity="{p["blossom_op"]}" fill="{p["blossom"]}">'
        '<circle cx="180" cy="700" r="7"/><circle cx="205" cy="690" r="6"/><circle cx="195" cy="712" r="6"/>'
        '<circle cx="1260" cy="680" r="7"/><circle cx="1285" cy="695" r="6"/><circle cx="1270" cy="668" r="5"/>'
        '</g></svg>'
    )

_BG_PALETTES = {
    # Sáng (5h–11h): bình minh hồng cam, mặt trời thấp ấm áp, hồ ánh cam-xanh
    "morning": dict(sky0="#bfe3f5", sky1="#ffd9c2", sky2="#ffc9a3",
                    orb="#ffb56b", orb_x=1080, orb_y=380, orb_r=90, orb_op="0.85",
                    mtn0="#8fa8c8", mtn1="#7089ad", far0="#e0c4d0", far1="#c9aec2",
                    snow="#fff6ee", mist_op="0.45",
                    water0="#ffd0a8", water1="#8fb0d0",
                    blossom="#f2a9c4", blossom_op="0.55"),
    # Trưa/chiều (11h–17h): trời xanh trong, nắng cao rực rỡ, hồ xanh biếc
    "noon": dict(sky0="#9fd4f5", sky1="#c9e8fa", sky2="#eaf6fd",
                 orb="#fff3b0", orb_x=1080, orb_y=160, orb_r=80, orb_op="0.9",
                 mtn0="#7f9cc0", mtn1="#5f7fa8", far0="#b8cfe6", far1="#9fbdd9",
                 snow="#fbfdff", mist_op="0.45",
                 water0="#bfe0f5", water1="#7fb4dd",
                 blossom="#f2c9d8", blossom_op="0.5"),
    # Tối (17h–5h): trời tím đậm, trăng vàng và sao, hồ đêm thẫm — đồng bộ với splash đêm
    "night": dict(sky0="#16213e", sky1="#3a3f66", sky2="#6b5b8a",
                  orb="#ffe9a8", orb_x=1080, orb_y=200, orb_r=70, orb_op="0.9",
                  mtn0="#3d4a6b", mtn1="#2c3854", far0="#4a5680", far1="#3d4a6b",
                  snow="#dfe6f5", mist_op="0.16",
                  water0="#3b3f66", water1="#1d2440",
                  blossom="#a86f8a", blossom_op="0.35", stars=True),
}

@st.cache_resource
def _build_bg_switcher_js():
    """Sinh 3 cảnh nền + đóng gói JS — cache để không chạy lại ở mỗi rerun."""
    b64 = {k: _b64.b64encode(_fuji_svg(v).encode()).decode() for k, v in _BG_PALETTES.items()}
    js = """
<script>
(function(){
    var doc = window.parent.document;
    var BG = {
        morning: "__B64_MORNING__",
        noon: "__B64_NOON__",
        night: "__B64_NIGHT__"
    };
    function pick(){
        var h = new Date().getHours();
        if (h >= 5 && h < 11) return 'morning';
        if (h >= 11 && h < 17) return 'noon';
        return 'night';
    }
    function apply(){
        var app = doc.querySelector('.stApp');
        if (!app) { setTimeout(apply, 300); return; }
        // Dọn splash bị kẹt (nếu có) — chạy mỗi lần rerun nên luôn tự sửa
        var sp = doc.getElementById('splash-overlay');
        if (sp && !sp.__born) { sp.__born = Date.now(); }
        if (sp && Date.now() - sp.__born > 8000) { sp.remove(); }
        var k = pick();
        // Trạng thái lưu trên DOM (không phải biến iframe) → rerun không re-apply, không flash
        if (app.dataset.hpBg === k) return;
        app.dataset.hpBg = k;
        app.style.transition = 'background 1.2s ease';
        app.style.background = 'url("data:image/svg+xml;base64,' + BG[k] + '") center 35% / cover no-repeat fixed';
    }
    apply();
    setInterval(apply, 60000); // tự chuyển cảnh khi bước sang khung giờ khác
})();
</script>
"""
    return (js.replace("__B64_MORNING__", b64["morning"])
              .replace("__B64_NOON__", b64["noon"])
              .replace("__B64_NIGHT__", b64["night"]))

components.html(_build_bg_switcher_js(), height=0)

components.html("""
<div style="
    display:flex; align-items:center; justify-content:space-between;
    flex-wrap:wrap; gap:8px;
    font-family:'Source Sans Pro','Segoe UI',Arial,sans-serif;
    padding-bottom:1.5rem; border-bottom:1px solid #ececec;
">
    <div style="font-size:1.5rem; font-weight:650; color:#1a1a1a; letter-spacing:-0.02em;">
        <span id="greet-emoji" style="margin-right:6px;"></span><span id="greet-text"></span>, Ta Chau
    </div>
    <div id="date-wrap" style="font-size:0.85rem; color:#9b9b9b; font-weight:500; font-variant-numeric:tabular-nums;">
        <span id="date-text"></span> · <span id="live-clock" style="color:#6b6b6b; font-weight:600;"></span>
    </div>
</div>
<script>
    var WD = ["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"];
    function pad(n){ return String(n).padStart(2,'0'); }
    function update(){
        var d = new Date();
        var h = d.getHours();
        // Time slots: 0-12 morning, 12-18 afternoon, 18-24 evening
        var greet, emoji;
        if (h < 12) { greet = "Good morning"; emoji = "☀️"; }
        else if (h < 18) { greet = "Good afternoon"; emoji = "🌆"; }
        else { greet = "Good evening"; emoji = "🌙"; }
        var ge = document.getElementById('greet-emoji');
        var gt = document.getElementById('greet-text');
        var dt = document.getElementById('date-text');
        var lc = document.getElementById('live-clock');
        if (ge) ge.textContent = emoji;
        if (gt) gt.textContent = greet;
        if (dt) dt.textContent = WD[d.getDay()] + ", " + pad(d.getDate()) + "/" + pad(d.getMonth()+1) + "/" + d.getFullYear();
        if (lc) lc.textContent = pad(h) + ":" + pad(d.getMinutes()) + ":" + pad(d.getSeconds());
        // Khung nội dung nền trắng đục → chữ đậm cố định, dễ đọc mọi khung giờ
        var wrap = document.getElementById('date-wrap');
        if (wrap && lc) {
            wrap.style.color = '#5a5a6e';
            lc.style.color = '#33334a';
        }
    }
    setInterval(update, 1000);
    update();
</script>
""", height=64)

# Hiệu ứng trang trí: hoa anh đào rơi + mây trôi — chèn NGOÀI khung (sau khung che)
components.html("""
<script>
(function(){
    var doc = window.parent.document;
    // Nếu lớp trang trí đã tồn tại thì GIỮ NGUYÊN (tránh mây/chim giật lại từ đầu mỗi rerun)
    if (doc.getElementById('deco-layer-outer')) return;

    // CSS
    var css = doc.createElement('style');
    css.id = 'deco-style';
    css.textContent = `
      #deco-layer-outer{position:fixed;inset:0;pointer-events:none;z-index:0;overflow:hidden;}
      #deco-layer-outer .cloud{position:fixed;font-size:2.8rem;opacity:.5;animation:drift linear infinite;}
      @keyframes drift{from{transform:translateX(-140px);}to{transform:translateX(112vw);}}
      #deco-layer-outer .petal{position:fixed;top:-40px;font-size:1.1rem;animation:fall linear infinite;}
      @keyframes fall{0%{transform:translateY(-40px) translateX(0) rotate(0);opacity:0;}12%{opacity:.85;}100%{transform:translateY(106vh) translateX(50px) rotate(360deg);opacity:.65;}}
      #deco-layer-outer .bird{position:fixed;opacity:.6;animation:flyby linear infinite;}
      @keyframes flyby{from{transform:translateX(-90px) translateY(0);}50%{transform:translateX(50vw) translateY(-22px);}to{transform:translateX(112vw) translateY(8px);}}
      #deco-layer-outer .bird svg{overflow:visible;display:block;animation:bob 1.1s ease-in-out infinite alternate;}
      @keyframes bob{from{transform:translateY(-3px);}to{transform:translateY(3px);}}
      #deco-layer-outer .bwing{transform-box:fill-box;transform-origin:88% 92%;
        animation:flap .42s ease-in-out infinite alternate;}
      #deco-layer-outer .bwing.back{animation-delay:.21s;}
      @keyframes flap{from{transform:rotate(-38deg);}to{transform:rotate(26deg);}}
    `;
    doc.head.appendChild(css);

    // Chim SVG vỗ cánh (thân + đầu + mỏ + 2 cánh lệch pha)
    function birdSvg(w){
        return '<svg viewBox="0 0 64 44" width="' + w + '" xmlns="http://www.w3.org/2000/svg">' +
            '<path class="bwing back" d="M30 23 Q 40 4 54 8 Q 42 16 31 27 Z" fill="#e8b3c0"/>' +
            '<path d="M10 24 L23 20 L23 28 Z" fill="#c76f88"/>' +
            '<ellipse cx="33" cy="24" rx="12" ry="8.5" fill="#d98a9e"/>' +
            '<circle cx="46" cy="17" r="6.5" fill="#d98a9e"/>' +
            '<path d="M52 15 L59 17.5 L52 20 Z" fill="#f0a25f"/>' +
            '<circle cx="48" cy="16" r="1.4" fill="#4a2f38"/>' +
            '<path class="bwing" d="M33 22 Q 22 2 8 6 Q 20 15 34 26 Z" fill="#c76f88"/>' +
            '</svg>';
    }

    // Lớp trang trí
    var d = doc.createElement('div');
    d.id = 'deco-layer-outer';
    d.innerHTML = `
      <div class="cloud" style="top:7%; left:0; animation-duration:52s;">☁️</div>
      <div class="cloud" style="top:15%; left:0; animation-duration:70s; animation-delay:9s; font-size:2rem;">☁️</div>
      <div class="cloud" style="top:4%; left:0; animation-duration:84s; animation-delay:22s; font-size:2.3rem; opacity:.4;">☁️</div>
      <div class="bird" style="top:21%; animation-duration:38s; animation-delay:6s;">` + birdSvg(44) + `</div>
      <div class="bird" style="top:25%; animation-duration:46s; animation-delay:20s; opacity:.45;">` + birdSvg(30) + `</div>
      <div class="bird" style="top:12%; animation-duration:42s; animation-delay:33s; opacity:.5;">` + birdSvg(36) + `</div>
      <div class="petal" style="left:3%; animation-duration:12s; animation-delay:0s;">🌸</div>
      <div class="petal" style="left:8%; animation-duration:15s; animation-delay:3s;">🌸</div>
      <div class="petal" style="left:13%; animation-duration:10s; animation-delay:6s; font-size:.9rem;">🌸</div>
      <div class="petal" style="left:5%; animation-duration:17s; animation-delay:9s; font-size:.8rem;">🌸</div>
      <div class="petal" style="left:83%; animation-duration:13s; animation-delay:1s;">🌸</div>
      <div class="petal" style="left:89%; animation-duration:16s; animation-delay:4s;">🌸</div>
      <div class="petal" style="left:94%; animation-duration:11s; animation-delay:7s; font-size:.9rem;">🌸</div>
      <div class="petal" style="left:86%; animation-duration:18s; animation-delay:10s; font-size:.8rem;">🌸</div>
    `;
    // Chèn làm CON ĐẦU TIÊN của .stApp → nằm trên nền Phú Sĩ, dưới khung nội dung (khung đặc che)
    var app = doc.querySelector('.stApp') || doc.body;
    app.insertBefore(d, app.firstChild);
})();
</script>
""", height=0)

# Menu selection (session state)
if "menu" not in st.session_state:
    st.session_state.menu = None

def go_menu(name):
    st.session_state.menu = name

# ── Menu landing screen ───────────────────────────────────────────────────

# ── Tạo file ARR từ file Arrival (Book) Smile ──────────────────────────────
def build_arr(book_bytes):
    """Tạo file ARR định dạng in chuẩn (theo file mẫu arr_09_07):
    - Nhóm theo Conf#: mỗi booking 1 dòng cao 126, Conf# chữ to 45 đậm nền xanh,
      số phòng (J) chữ 40 đậm, viền thin toàn bộ, căn giữa, Times New Roman.
    - Dòng phụ CÀ THẺ / THU TIỀN / LUNCH IN: cao 43.5, gộp C:I chữ 35 đậm nền xanh,
      ô J gộp dọc từ dòng booking xuống hết dòng phụ.
    - Cột A, G ẩn; ngày định dạng mm-dd-yy; in dọc scale 64%.

    Quy tắc dữ liệu:
    - Conf# trùng nhau = 1 booking; J = số phòng UNIQUE (loại phòng ảo 9000-9999).
    - CÀ THẺ: Company AGODA/EXPEDIA/CTRIP · THU TIỀN: BOOKING.COM hoặc Notice chứa
      "RC PAY BF C/I" (KHÔNG tính "RC TA PAY BF C/I") · LUNCH IN: Notice chứa "LUNCH".
    """
    import re as _re2
    import numpy as _np
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.styles.colors import Color

    df = pd.read_excel(io.BytesIO(book_bytes), header=None)
    hdr = None
    for i in range(min(5, len(df))):
        if any(str(v).strip() == 'Conf#' for v in df.iloc[i] if pd.notna(v)):
            hdr = i; break
    data = df.iloc[(hdr + 1 if hdr is not None else 0):].copy()
    if data.shape[1] < 47:
        raise ValueError("File không đúng cấu trúc Arrival Smile (thiếu cột). Vui lòng kiểm tra lại file.")
    data['conf'] = data[5].ffill()
    data = data[data['conf'].notna()]
    if len(data) == 0:
        raise ValueError("File không có dữ liệu booking nào.")

    CA_THE = ('agoda', 'expedia', 'ctrip')

    def _num_or_keep(v):
        """int nếu là số, datetime giữ nguyên, còn lại trả chuỗi/None."""
        import datetime as _dt
        if v is None or (not isinstance(v, str) and pd.isna(v)): return None
        if isinstance(v, pd.Timestamp): return v.to_pydatetime()
        if isinstance(v, _dt.datetime): return v
        if isinstance(v, (_np.floating, float)):
            f = float(v); return int(f) if f == int(f) else f
        if isinstance(v, (_np.integer, int)): return int(v)
        s = str(v).strip()
        try:
            f = float(s); return int(f) if f == int(f) else f
        except Exception:
            return s if s else None

    # ── Style theo file mẫu ──
    _thin = Side(style='thin')
    BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    GREEN = PatternFill(fill_type='solid', fgColor=Color(theme=9, tint=0.7999816888943144))
    F_BASE = Font(name='Times New Roman', size=17)
    F_CONF = Font(name='Times New Roman', size=45, bold=True)
    F_ROOM = Font(name='Times New Roman', size=40, bold=True)
    F_EXTRA = Font(name='Times New Roman', size=35, bold=True)
    DATE_FMT = 'mm-dd-yy'

    wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
    er = 0
    stats = {'bookings': 0, 'ca_the': 0, 'thu_tien': 0, 'lunch': 0}

    def _style_row(r):
        for ci in range(1, 11):
            c = ws.cell(r, ci)
            c.border = BORDER; c.alignment = CENTER
            if c.font.name != 'Times New Roman': c.font = F_BASE

    for _conf, grp in data.groupby('conf', sort=False):
        first = grp.iloc[0]
        rooms = grp[10].dropna().astype(str).str.strip()
        rooms = [r for r in rooms.unique() if r and not _re2.fullmatch(r'9\d{3}', r)]
        notice = str(first[46]).strip() if pd.notna(first[46]) else ''
        company = str(first[19]).strip() if pd.notna(first[19]) else ''

        # ── Dòng booking ──
        er += 1
        book_r = er
        ws.row_dimensions[er].height = 126.0
        _style_row(er)
        vals = {1: _num_or_keep(first[0]), 3: _num_or_keep(first[5]),
                4: _num_or_keep(first[13]), 5: _num_or_keep(first[14]),
                6: company or None, 7: _num_or_keep(first[20]),
                9: notice or None, 10: len(rooms)}
        for ci, v in vals.items():
            c = ws.cell(er, ci)
            if v is not None: c.value = v
            if hasattr(v, 'year'): c.number_format = DATE_FMT
        ws.cell(er, 3).font = F_CONF; ws.cell(er, 3).fill = GREEN
        ws.cell(er, 10).font = F_ROOM
        stats['bookings'] += 1

        # ── Dòng phụ ──
        comp_key = _norm_nat(company)
        labels = []
        if any(comp_key.startswith(x) for x in CA_THE):
            labels.append('CÀ THẺ'); stats['ca_the'] += 1
        elif comp_key.startswith('booking'):
            labels.append('THU TIỀN'); stats['thu_tien'] += 1
        if 'RC PAY BF C/I' in notice.upper() and 'THU TIỀN' not in labels:
            labels.append('THU TIỀN'); stats['thu_tien'] += 1
        if 'LUNCH' in notice.upper():
            labels.append('LUNCH IN'); stats['lunch'] += 1

        extra_c0 = _num_or_keep(grp.iloc[1][0]) if len(grp) > 1 else None
        for lb in labels:
            er += 1
            ws.row_dimensions[er].height = 43.5
            _style_row(er)
            if extra_c0 is not None: ws.cell(er, 1).value = extra_c0
            ws.merge_cells(start_row=er, start_column=3, end_row=er, end_column=9)
            c = ws.cell(er, 3); c.value = lb
            c.font = F_EXTRA; c.fill = GREEN
        if labels:
            # Gộp ô J (số phòng) từ dòng booking xuống hết dòng phụ
            ws.merge_cells(start_row=book_r, start_column=10, end_row=er, end_column=10)

    # ── Cột: độ rộng + ẩn A, G ──
    from openpyxl.utils import get_column_letter
    widths = {'A': 0.0, 'C': 33.43, 'D': 15.29, 'E': 14.71, 'F': 20.71,
              'G': 0.0, 'I': 50.71, 'J': 10.29}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
        if w == 0.0: ws.column_dimensions[col].hidden = True

    # ── Thiết lập in như file mẫu ──
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.scale = 64
    ws.page_margins = PageMargins(left=0.7, right=0.45, top=0.0, bottom=0.0, header=0.3, footer=0.3)
    ws.sheet_view.view = 'pageBreakPreview'

    return wb, stats



if st.session_state.menu is None:
    st.markdown('<div class="section-label">✨ Chọn chức năng</div>', unsafe_allow_html=True)
    st.write("")
    mcol1, mcol2, mcol3 = st.columns(3)
    with mcol1:
        st.markdown("""
        <div class="menu-card menu-blue">
            <div class="menu-icon">📋</div>
            <div class="menu-title">Xử lý hồ sơ hàng ngày</div>
            <div class="menu-desc">Quy đổi tỷ giá · Tách Quốc tế/Việt Nam · KBTT · Thông báo lưu trú · ĐK14</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Mở  →", key="btn_daily", use_container_width=True,
                  on_click=go_menu, args=("daily",))
    with mcol2:
        st.markdown("""
        <div class="menu-card menu-green">
            <div class="menu-icon">🖨️</div>
            <div class="menu-title">Tạo Regcard PDF</div>
            <div class="menu-desc">Điền dữ liệu booking lên mẫu Registration Card · Xuất PDF hàng loạt + file ARR</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Mở  →", key="btn_regcard", use_container_width=True,
                  on_click=go_menu, args=("regcard",))
    with mcol3:
        st.markdown("""
        <div class="menu-card menu-amber">
            <div class="menu-icon">🔍</div>
            <div class="menu-title">Đối chiếu lưu trú</div>
            <div class="menu-desc">So khớp Smile vs Trang quản lý người nước ngoài · Tìm thiếu/trùng/lệch phòng</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Mở  →", key="btn_recon", use_container_width=True,
                  on_click=go_menu, args=("recon",))

# ── Daily processing screen ───────────────────────────────────────────────
if st.session_state.menu == "daily":
    st.button("←  Quay lại menu", key="back_daily", on_click=go_menu, args=(None,))
    st.write("")
    st.markdown('<div class="section-label">⚙️ Cài đặt</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        rate = st.number_input("💱 Tỷ giá USD/EUR → VNĐ", value=29535.15, step=0.01, format="%.2f")
    with col2:
        today = datetime.date.today()
        date_str = st.text_input("📅 Ngày (dùng cho tên file)", value=f"{today.day}_{today.month:02d}")

    st.write("")
    st.markdown('<div class="section-label">📂 Tải file lên</div>', unsafe_allow_html=True)
    col_x, col_s = st.columns(2)
    with col_x:
        xlsx_file = st.file_uploader("File XLSX — Dữ liệu khách (bắt buộc)", type=['xlsx'], key="daily_xlsx")
    with col_s:
        xls_file = st.file_uploader("File XLS — Nguồn ĐK14 (tùy chọn)", type=['xls'], key="daily_xls")

    st.write("")

    if st.button("⚡ Bắt đầu xử lý", type="primary", disabled=(xlsx_file is None and xls_file is None), use_container_width=True):
        with st.spinner("Đang xử lý..."):
            try:
                progress = st.progress(0, text="Bắt đầu...")
                zip_buf = io.BytesIO()
                files_made = []
                has_xlsx = xlsx_file is not None
                has_dk14 = False
                conv = 0; gks_cnt = 0; gbl_cnt = 0
                df = None; df_intl = None; df_vn = None

                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    # ── Xử lý file XLSX (nếu có) ──
                    if has_xlsx:
                        progress.progress(10, text="Quy đổi tỷ giá...")
                        xlsx_bytes = xlsx_file.read()
                        wb, conv = process_xlsx(xlsx_bytes, rate)
                        df = pd.read_excel(io.BytesIO(xlsx_bytes))
                        df_intl = df[df['LOẠI KHÁCH']=='Quốc tế'].reset_index(drop=True)
                        df_vn   = df[df['LOẠI KHÁCH']=='Việt Nam'].reset_index(drop=True)

                        progress.progress(30, text="Tách file Quốc tế / Việt Nam...")
                        wb_intl = split_wb(wb, 'Quốc tế')
                        wb_vn   = split_wb(wb, 'Việt Nam')

                        progress.progress(45, text="Điền mẫu KBTT...")
                        wb_kbtt = build_kbtt(df_intl)

                        progress.progress(60, text="Điền mẫu Thông báo lưu trú VNM...")
                        wb_vnm, gks_cnt, gbl_cnt = build_vnm(df_vn)

                        zf.writestr(f'converted_{date_str}.xlsx',      wb_to_bytes(wb))
                        zf.writestr(f'KhachQuocTe_{date_str}.xlsx',    wb_to_bytes(wb_intl))
                        zf.writestr(f'KhachVietNam_{date_str}.xlsx',   wb_to_bytes(wb_vn))
                        zf.writestr(f'ho_so_KBTT_NNN_{date_str}.xlsx', wb_to_bytes(wb_kbtt))
                        zf.writestr(f'thong_bao_luu_tru_VNM_{date_str}.xlsx', wb_to_bytes(wb_vnm))
                        files_made += ["📄 converted (file chung)", "🌍 KhachQuocTe", "🇻🇳 KhachVietNam",
                                       "📝 KBTT NNN", "📑 Thông báo lưu trú VNM"]

                    # ── Xử lý file ĐK14 (độc lập, chỉ cần file XLS) ──
                    if xls_file:
                        progress.progress(85, text="Điền mẫu ĐK14...")
                        xls_bytes = xls_file.read()
                        wb_dk14, dk_count = build_dk14(xls_bytes)
                        zf.writestr(f'dk14_{date_str}.xlsx', wb_to_bytes(wb_dk14))
                        has_dk14 = True
                        files_made.append("🚔 ĐK14")

                progress.progress(100, text="Hoàn tất!")
                progress.empty()

                # Lưu kết quả vào session — kết quả & nút tải không biến mất sau rerun
                _daily = {'files_made': files_made, 'zip': zip_buf.getvalue(),
                          'date_str': date_str, 'has_xlsx': has_xlsx, 'has_dk14': has_dk14}
                if has_xlsx:
                    unknown_nats = []
                    for q in df_intl.get('QUỐC TỊCH', pd.Series([], dtype=str)).dropna().unique():
                        mapped = lookup_nat_kbtt(q)
                        if not _re.match(r'^[A-Z]{2,3} - ', str(mapped)):
                            unknown_nats.append(str(q))
                    _daily.update({'total': len(df), 'intl': len(df_intl), 'vn': len(df_vn),
                                   'gks': gks_cnt, 'gbl': gbl_cnt, 'conv': conv,
                                   'unknown_nats': unknown_nats})
                st.session_state['daily_results'] = _daily
            except Exception as e:
                st.session_state.pop('daily_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    _dr = st.session_state.get('daily_results')
    if _dr:
        st.success("✅ Xử lý hoàn tất!")
        if _dr['has_xlsx']:
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Tổng khách", _dr['total'])
            c2.metric("Quốc tế", _dr['intl'])
            c3.metric("Việt Nam", _dr['vn'])
            c4.metric("GKS + GBL", f"{_dr['gks']} + {_dr['gbl']}")
            st.info(f"💱 Đã quy đổi tỷ giá cho **{_dr['conv']}** ô (đã tô vàng)")
            if _dr['unknown_nats']:
                st.warning("⚠️ Quốc tịch chưa có mã (giữ nguyên tên, cần kiểm tra): " + ", ".join(_dr['unknown_nats']))
        elif _dr['has_dk14']:
            st.info("ℹ️ Chỉ tạo file ĐK14 (không có file XLSX dữ liệu khách).")

        st.markdown("**File đã tạo:** " + " · ".join(_dr['files_made']))

        st.download_button(
            label="⬇️ Tải về tất cả file (ZIP)",
            data=_dr['zip'],
            file_name=f"hotel_{_dr['date_str']}.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )

# ── Regcard screen ────────────────────────────────────────────────────────
if st.session_state.menu == "regcard":
    st.button("←  Quay lại menu", key="back_regcard", on_click=go_menu, args=(None,))
    st.write("")
    st.markdown('<div class="section-label">🖨️ Tạo Registration Card + file ARR</div>', unsafe_allow_html=True)
    st.caption("Điền dữ liệu từ file Arrival Smile lên mẫu Regcard PDF gốc, đồng thời tạo file ARR "
               "(nhóm Conf# · đếm phòng · Cà Thẻ / Thu Tiền / Lunch In) — ra cùng lúc 2 file.")

    rc_file = st.file_uploader("File Excel dữ liệu booking (.xlsx)", type=['xlsx'], key="rc_xlsx")

    only_main = st.checkbox("Chỉ tạo cho khách chính (có mã Conf#)", value=True,
                            help="Bỏ chọn để tạo regcard cho tất cả khách, kể cả khách đi cùng phòng")

    st.write("")

    if st.button("🖨️ Tạo Regcard PDF + file ARR", type="primary", disabled=rc_file is None, use_container_width=True):
        with st.spinner("Đang tạo Regcard PDF và file ARR..."):
            try:
                rc_bytes = rc_file.read()
                pdf_data, count = build_regcards(rc_bytes, only_main=only_main)

                # Tạo file ARR từ cùng file đầu vào (lỗi ARR không làm hỏng PDF)
                arr_bytes, arr_stats, arr_err = None, None, None
                try:
                    wb_arr, arr_stats = build_arr(rc_bytes)
                    arr_bytes = wb_to_bytes(wb_arr)
                except Exception as _e:
                    arr_err = str(_e)

                st.session_state['rc_results'] = {
                    'pdf': pdf_data, 'count': count,
                    'arr': arr_bytes, 'arr_stats': arr_stats, 'arr_err': arr_err,
                    'date': datetime.date.today().strftime('%d_%m'),
                }
            except Exception as e:
                st.session_state.pop('rc_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    # Kết quả lưu trong session — 2 nút tải không biến mất sau khi bấm 1 nút
    _res = st.session_state.get('rc_results')
    if _res:
        if _res['count'] == 0:
            st.warning("⚠️ Không tìm thấy khách nào để tạo regcard. Kiểm tra lại file.")
        else:
            st.success(f"✅ Đã tạo {_res['count']} regcard"
                       + (" + file ARR!" if _res['arr'] else "!"))
            if _res['arr_stats']:
                a1, a2, a3, a4, a5 = st.columns(5)
                a1.metric("🖨️ Regcard", _res['count'])
                a2.metric("📦 Booking", _res['arr_stats']['bookings'])
                a3.metric("💳 Cà Thẻ", _res['arr_stats']['ca_the'])
                a4.metric("💵 Thu Tiền", _res['arr_stats']['thu_tien'])
                a5.metric("🍽️ Lunch In", _res['arr_stats']['lunch'])
            else:
                st.metric("Số regcard", _res['count'])

            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    label=f"⬇️ Tải {_res['count']} Regcard (PDF)",
                    data=_res['pdf'],
                    file_name=f"regcards_{_res['date']}.pdf",
                    mime="application/pdf",
                    use_container_width=True, type="primary", key="dl_rc_pdf")
            with d2:
                if _res['arr']:
                    st.download_button(
                        label="⬇️ Tải file ARR (Excel)",
                        data=_res['arr'],
                        file_name=f"arr_{_res['date']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary", key="dl_rc_arr")
            if _res['arr_err']:
                st.warning(f"⚠️ Không tạo được file ARR: {_res['arr_err']}")


# ── Đối chiếu: sub-menu 2 lựa chọn (có cổng mật khẩu riêng) ────────────────
RECON_PASS = "368736"
if "recon_ok" not in st.session_state:
    st.session_state.recon_ok = False

def _check_recon():
    if st.session_state.get("recon_pass_input", "") == RECON_PASS:
        st.session_state.recon_ok = True
        st.session_state.recon_pass_err = False
    else:
        st.session_state.recon_pass_err = True

if st.session_state.menu == "recon":
    st.button("←  Quay lại menu", key="back_recon", on_click=go_menu, args=(None,))
    st.write("")

    # Cổng mật khẩu cho tính năng đối chiếu
    if not st.session_state.recon_ok:
        st.markdown('<div class="section-label">🔒 Nhập mật khẩu để truy cập</div>', unsafe_allow_html=True)
        st.caption("Tính năng Đối chiếu lưu trú được bảo vệ bằng mật khẩu riêng.")
        with st.form("recon_pass_form", clear_on_submit=False):
            st.text_input("Mật khẩu", key="recon_pass_input", type="password", placeholder="Nhập mật khẩu")
            ok = st.form_submit_button("Mở khóa →", type="primary", use_container_width=True)
            if ok:
                _check_recon()
        if st.session_state.get("recon_pass_err"):
            st.error("❌ Mật khẩu không đúng!")
        st.stop()

    st.markdown('<div class="section-label">🔍 Chọn loại kiểm tra</div>', unsafe_allow_html=True)
    st.write("")
    scol1, scol2 = st.columns(2)
    with scol1:
        st.markdown("""
        <div class="menu-card menu-amber">
            <div class="menu-icon">🌏</div>
            <div class="menu-title">Kiểm tra lưu trú người nước ngoài</div>
            <div class="menu-desc">So khớp khách Smile vs Trang quản lý người nước ngoài · Tìm khách chưa đăng ký / đăng ký trùng</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Mở  →", key="btn_recon_person", use_container_width=True,
                  on_click=go_menu, args=("recon_person",))
    with scol2:
        st.markdown("""
        <div class="menu-card menu-gray">
            <div class="menu-icon">🚪</div>
            <div class="menu-title">Kiểm tra hệ thống quản lý lưu trú phòng</div>
            <div class="menu-desc">So khớp số phòng inhouse trên Smile với hệ thống quản lý lưu trú phòng · Tìm phòng chưa đăng ký / thừa / trùng</div>
        </div>
        """, unsafe_allow_html=True)
        st.button("Mở  →", key="btn_recon_room", use_container_width=True,
                  on_click=go_menu, args=("recon_room",))

# ── Kiểm tra lưu trú người nước ngoài ─────────────────────────────────────
if st.session_state.menu == "recon_person":
    st.button("←  Quay lại", key="back_recon_person", on_click=go_menu, args=("recon",))
    st.write("")
    st.markdown('<div class="section-label">🌏 Kiểm tra lưu trú người nước ngoài</div>', unsafe_allow_html=True)
    st.caption("So khớp khách Smile vs Trang quản lý người nước ngoài theo số hộ chiếu — tìm khách chưa đăng ký / đăng ký trùng.")

    rc1, rc2 = st.columns(2)
    with rc1:
        smile_file = st.file_uploader("File Smile — Inhouse (.xlsx)", type=['xlsx'], key="recon_smile")
    with rc2:
        luutru_file = st.file_uploader("File Trang lưu trú người nước ngoài (.xlsx)", type=['xlsx'], key="recon_luutru")

    today_str = st.text_input("📅 Ngày xuất file (hôm nay)", value=datetime.date.today().strftime('%d/%m/%Y'),
                              help="Dùng để loại bỏ: khách arrival hôm nay (Smile) và khách ngày đi dự kiến hôm nay (Lưu trú)")

    st.write("")

    if st.button("🔍 Bắt đầu kiểm tra", type="primary",
                 disabled=(smile_file is None or luutru_file is None), use_container_width=True):
        with st.spinner("Đang đối chiếu..."):
            try:
                today = pd.to_datetime(today_str, format='%d/%m/%Y')
                st.session_state['recon_results'] = reconcile(smile_file.read(), luutru_file.read(), today)
            except Exception as e:
                st.session_state.pop('recon_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    r = st.session_state.get('recon_results')
    if r:
        st.success("✅ Kiểm tra hoàn tất!")

        c1, c2 = st.columns(2)
        c1.metric("Smile (sau lọc)", r['smile_filtered'], f"từ {r['smile_total']} (bỏ VNM + arrival/departure hôm nay)")
        c2.metric("Lưu trú (sau lọc)", r['luutru_filtered'], f"từ {r['luutru_total']} (bỏ đi dự kiến hôm nay)")

        st.divider()
        st.markdown("### 👤 Kết quả đối chiếu người (theo số hộ chiếu)")

        n_chua = len(r['chua_dk']); n_thua = len(r['thua']); n_dup = len(r['dup'])

        # Bảng tổng hợp chênh lệch
        m1, m2, m3 = st.columns(3)
        m1.metric("🔴 Chưa đăng ký", n_chua)
        m2.metric("🟡 Có/lưu trú, thiếu/Smile", n_thua)
        m3.metric("🟠 Đăng ký trùng", n_dup)

        if n_chua == 0 and n_thua == 0 and n_dup == 0:
            st.success("✅ Khớp hoàn toàn! Không có ai thiếu/thừa/trùng.")

        if n_chua > 0:
            st.error(f"🔴 {n_chua} khách trên Smile nhưng CHƯA đăng ký lưu trú:")
            st.dataframe(r['chua_dk'], use_container_width=True, hide_index=True)
        else:
            st.success("✅ Không có khách nào chưa đăng ký lưu trú.")

        if n_thua > 0:
            st.warning(f"🟡 Chênh lệch **{n_thua} người**: có trên Trang quản lý người nước ngoài nhưng KHÔNG có trên Smile (có thể đã checkout nhưng chưa xóa khỏi lưu trú):")
            st.dataframe(r['thua'], use_container_width=True, hide_index=True)
            # Nút tải danh sách chênh lệch
            _csv = r['thua'].to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Tải danh sách chênh lệch (CSV)", _csv,
                               file_name="chenh_lech_luu_tru.csv", mime="text/csv")

        if n_dup > 0:
            st.warning(f"🟠 {n_dup} dòng ĐĂNG KÝ TRÙNG trên lưu trú:")
            st.dataframe(r['dup'], use_container_width=True, hide_index=True)

# ── Kiểm tra hệ thống quản lý lưu trú phòng ────────────────────────────────
def reconcile_rooms(smile_bytes, room_bytes, today):
    """Đối chiếu phòng inhouse từ file khách lưu trú Smile (trừ khách Arrival hôm nay)
    với file Excel chỉ chứa danh sách số phòng."""
    from collections import Counter

    # ── Đọc file khách lưu trú Smile ──
    df1 = pd.read_excel(io.BytesIO(smile_bytes), header=0)
    if 'Rm#' not in df1.columns:
        raise ValueError("File Smile không có cột 'Rm#'. Vui lòng dùng file khách lưu trú xuất từ Smile.")
    smile = df1.dropna(subset=['Rm#']).copy()
    smile['room'] = smile['Rm#'].apply(_norm_room)
    smile['Arrival'] = pd.to_datetime(smile['Arrival'], errors='coerce') if 'Arrival' in smile else pd.NaT
    _ln = smile['Last Name'].astype(str).str.strip() if 'Last Name' in smile else ''
    _fn = smile['First Name'].astype(str).str.strip() if 'First Name' in smile else ''
    smile['name'] = (_ln + ' ' + _fn).str.strip() if 'Last Name' in smile else ''
    smile_total = len(smile)
    # Cột Departure (dò tên linh hoạt)
    _dep_col = next((c for c in df1.columns if 'depart' in str(c).lower()), None)
    smile['Departure'] = pd.to_datetime(df1[_dep_col], errors='coerce') if _dep_col else pd.NaT
    # Trừ khách Arrival = hôm nay và khách Departure = hôm nay (trả phòng)
    if 'Arrival' in smile:
        smile_f = smile[(smile['Arrival'].dt.date != today.date()) &
                        (smile['Departure'].dt.date != today.date())].copy()
    else:
        smile_f = smile.copy()

    # ── Đọc file chỉ chứa số phòng: lấy tất cả ô có dữ liệu ──
    raw = pd.read_excel(io.BytesIO(room_bytes), header=None, dtype=str)
    rooms_sys = []
    for _, row_vals in raw.iterrows():
        for v in row_vals:
            if pd.isna(v): continue
            r = _norm_room(v)
            if not r: continue
            # bỏ ô tiêu đề nếu lỡ có (vd "Số phòng", "Room", "STT")
            if _norm_nat(r) in ('sophong', 'phong', 'room', 'rm', 'stt'): continue
            rooms_sys.append(r)
    if not rooms_sys:
        raise ValueError("File số phòng không có dữ liệu. Vui lòng kiểm tra lại file.")

    sys_rooms = set(rooms_sys)
    _cnt = Counter(rooms_sys)
    sys_dup = sorted((r for r, c in _cnt.items() if c > 1), key=lambda x: (len(x), x))

    import re as _re3
    def _is_virtual(r):
        return bool(_re3.fullmatch(r'9\d{3}', r))  # phòng ảo 9000-9999 (posting master)

    smile_rooms = set(r for r in smile_f['room'] if r and not _is_virtual(r))
    sys_rooms = set(r for r in sys_rooms if not _is_virtual(r))

    def _sortkey(x): return (len(x), x)
    room_chua = sorted(smile_rooms - sys_rooms, key=_sortkey)  # inhouse nhưng CHƯA có trong file phòng
    room_thua = sorted(sys_rooms - smile_rooms, key=_sortkey)  # có trong file phòng nhưng KHÔNG còn inhouse

    # Chi tiết khách trong các phòng chưa đăng ký (tiện đăng ký bổ sung)
    if room_chua:
        detail = smile_f[smile_f['room'].isin(room_chua)].copy()
        detail['_arr'] = detail['Arrival'].dt.strftime('%d/%m/%Y') if 'Arrival' in detail else ''
        _nat = detail['NAT'] if 'NAT' in detail else ''
        detail = pd.DataFrame({
            'Số phòng': detail['room'].values,
            'Họ tên': detail['name'].values,
            'Quốc tịch': _nat.values if hasattr(_nat, 'values') else _nat,
            'Ngày đến': detail['_arr'].values if hasattr(detail['_arr'], 'values') else '',
        }).sort_values('Số phòng', key=lambda s: s.map(lambda x: (len(x), x)))
    else:
        detail = pd.DataFrame(columns=['Số phòng', 'Họ tên', 'Quốc tịch', 'Ngày đến'])

    return {
        'smile_total': smile_total, 'smile_filtered': len(smile_f),
        'sys_total': len(rooms_sys), 'sys_unique': len(sys_rooms),
        'smile_rooms': len(smile_rooms),
        'room_chua': room_chua, 'room_thua': room_thua, 'sys_dup': sys_dup,
        'room_match': len(smile_rooms & sys_rooms),
        'detail_chua': detail,
    }


if st.session_state.menu == "recon_room":
    st.button("←  Quay lại", key="back_recon_room", on_click=go_menu, args=("recon",))
    st.write("")
    st.markdown('<div class="section-label">🚪 Kiểm tra hệ thống quản lý lưu trú phòng</div>', unsafe_allow_html=True)
    st.caption("So khớp số phòng inhouse từ file khách lưu trú Smile với file danh sách số phòng — tìm phòng chưa đăng ký / thừa / trùng.")

    rr1, rr2 = st.columns(2)
    with rr1:
        smile_file_r = st.file_uploader("File khách lưu trú Smile (.xlsx)", type=['xlsx'], key="reconr_smile")
    with rr2:
        room_file = st.file_uploader("File số phòng (.xlsx — chỉ chứa danh sách số phòng)", type=['xlsx'], key="reconr_room")

    today_str_r = st.text_input("📅 Ngày xuất file (hôm nay)", value=datetime.date.today().strftime('%d/%m/%Y'),
                                key="reconr_today",
                                help="Khách có Arrival = ngày này trên Smile sẽ được loại bỏ khỏi đối chiếu")

    st.write("")

    if st.button("🔍 Bắt đầu kiểm tra", type="primary", key="reconr_run",
                 disabled=(smile_file_r is None or room_file is None), use_container_width=True):
        with st.spinner("Đang đối chiếu phòng..."):
            try:
                today_r = pd.to_datetime(today_str_r, format='%d/%m/%Y')
                st.session_state['reconr_results'] = reconcile_rooms(smile_file_r.read(), room_file.read(), today_r)
            except Exception as e:
                st.session_state.pop('reconr_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    rr = st.session_state.get('reconr_results')
    if rr:
        st.success("✅ Kiểm tra hoàn tất!")

        c1, c2, c3 = st.columns(3)
        c1.metric("Phòng inhouse (Smile)", rr['smile_rooms'],
                  f"{rr['smile_filtered']} khách (từ {rr['smile_total']}, đã trừ arrival hôm nay)")
        c2.metric("Phòng trong file", rr['sys_unique'],
                  (f"{rr['sys_total']} dòng" if rr['sys_total'] != rr['sys_unique'] else None))
        c3.metric("🟢 Phòng khớp", rr['room_match'])

        st.divider()
        st.markdown("### 🚪 Kết quả đối chiếu phòng")

        n_chua = len(rr['room_chua']); n_thua = len(rr['room_thua']); n_dup = len(rr['sys_dup'])

        m1, m2, m3 = st.columns(3)
        m1.metric("🔴 Chưa đăng ký", n_chua)
        m2.metric("🟡 Thừa trong file", n_thua)
        m3.metric("🟠 Trùng trong file", n_dup)

        if n_chua == 0 and n_thua == 0 and n_dup == 0:
            st.success("✅ Khớp hoàn toàn! Không có phòng thiếu/thừa/trùng.")
            st.balloons()

        if n_chua > 0:
            st.error(f"🔴 {n_chua} phòng có khách inhouse nhưng CHƯA có trong file số phòng: "
                     + ", ".join(rr['room_chua']))
            st.markdown("**Chi tiết khách trong các phòng chưa đăng ký:**")
            st.dataframe(rr['detail_chua'], use_container_width=True, hide_index=True)
            _csv_r = rr['detail_chua'].to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Tải danh sách phòng chưa đăng ký (CSV)", _csv_r,
                               file_name="phong_chua_dang_ky.csv", mime="text/csv")
        else:
            st.success("✅ Tất cả phòng inhouse đều đã có trong file số phòng.")

        if n_thua > 0:
            st.warning(f"🟡 {n_thua} phòng có trong file nhưng KHÔNG còn khách inhouse "
                       f"(có thể đã checkout nhưng chưa gỡ): "
                       + ", ".join(rr['room_thua']))

        if n_dup > 0:
            st.warning(f"🟠 {n_dup} phòng bị TRÙNG (xuất hiện nhiều lần) trong file số phòng: "
                       + ", ".join(rr['sys_dup']))


